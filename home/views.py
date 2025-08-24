from django import forms
from django.shortcuts import render
from .forms import ClassForm, PlatformForm, IssuingAuthForm, EntityForm, LicenseForm, FormsForm, TypeForm,InOutForm,  DocsForm, ShelfForm, DSCRenewalForm,UsageLogsForm, DSCEntityForm, InitiationForm, DeliveryCollectionForm, InternalDSCForm, ExternalDSCForm, EditUserForm,EmailTemplateForm,UnassignShelfForm,AssignShelfForm,BulkDSCUploadForm,BulkEntityUploadForm
from .models import DSC_class,Platform, IssuingAuth, Entity, LicensePeriod, Forms, Type,InOut, DSCMaster, Docs, Shelf, DSCRenewal,UsageLogs, DSCEntity, CustomUser,DSCExtraField,DSCExtraData,EmailTemplate,ShelfAssignment,RenewalExtraField,DSCRenewalHistory
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from .forms import LoginForm
from django.contrib.auth import login, authenticate, logout
from .forms import  LoginForm
from .decorators import role_required 
from django.contrib.auth import get_user_model
from .forms import UserForm
from django.db.models import F
from django.utils.timezone import now
from datetime import timedelta, date
from django.core.paginator import Paginator
import csv
import pandas as pd
from django.http import HttpResponse, FileResponse, HttpResponseRedirect
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.http import JsonResponse
from django.core.serializers.json import DjangoJSONEncoder
import json
import os
import shutil
import tempfile
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.management import call_command
import io
from io import BytesIO
import datetime
from openpyxl import Workbook
from django.apps import apps
import pytz
from datetime import datetime
import re
from django.db.models import OuterRef, Subquery
from django.db.models.functions import Coalesce
from itertools import chain
from django.contrib.auth.decorators import login_required, user_passes_test
# Create your views here.


def index(request):
    return render(request, 'index.html')

#Class
@login_required
@role_required('admin')
def dsc_class(request):
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Class added successfully!")  # Add success message
        else:
            messages.error(request, "There was an error adding the class.")  # Handle form errors

    form = ClassForm()
    class_form = {
        'form': form,
        'class_data': DSC_class.objects.all(),
    }
    return render(request, 'class.html', class_form)

@login_required
@role_required('admin')
def edit_class(request, class_id):
    dsc_class = get_object_or_404(DSC_class, class_id=class_id)

    if request.method == 'POST':
        form = ClassForm(request.POST, instance=dsc_class)
        if form.is_valid():
            form.save()
            messages.success(request, "Class updated successfully!")  # Success message
            return redirect('class')
        else:
            messages.error(request, "There was an error updating the class.")  # Error message

    form = ClassForm(instance=dsc_class)  # Prepopulate the form with existing data
    return render(request, 'class.html', {'form': form, 'class_data': DSC_class.objects.all()})

@login_required
@role_required('admin')
def delete_class(request, class_id):
    dsc_class = get_object_or_404(DSC_class, class_id=class_id)
    if request.method == "POST":  # Ensure deletion only occurs on POST
        dsc_class.delete()
        messages.success(request, "Class deleted successfully!")
        return redirect('class')

    # Render confirmation page if not a POST request
    return render(request, 'delete_confirmation.html', {'object': dsc_class})


# Add or List Platforms
@login_required
@role_required('admin')
def platform(request):
    if request.method == 'POST':
        form = PlatformForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Platform added successfully!")  # Success message
        else:
            messages.error(request, "There was an error adding the platform.")  # Error message

    form = PlatformForm()
    platform_form = {
        'form': form,
        'platform_data': Platform.objects.all(),
    }
    return render(request, 'platform.html', platform_form)

# Edit Platform
@login_required
@role_required('admin')
def edit_platform(request, platform_id):
    platform = get_object_or_404(Platform, platform_id=platform_id)

    if request.method == 'POST':
        form = PlatformForm(request.POST, instance=platform)
        if form.is_valid():
            form.save()
            messages.success(request, "Platform updated successfully!")  # Success message
            return redirect('platform')
        else:
            messages.error(request, "There was an error updating the platform.")  # Error message

    form = PlatformForm(instance=platform)
    return render(request, 'platform.html', {'form': form, 'platform_data': Platform.objects.all()})

# Delete Platform
@login_required
@role_required('admin')
def delete_platform(request, platform_id):
    platform = get_object_or_404(Platform, platform_id=platform_id)
    if request.method == 'POST':  # Ensure deletion occurs only on POST
        platform.delete()
        messages.success(request, "Platform deleted successfully!")  # Success message
        return redirect('platform')

    # Optional: Render a confirmation page if needed
    return render(request, 'delete_confirmation.html', {'object': platform})





# Add or List Issuing Authorities
@login_required
@role_required('admin')
def issuingauth(request):
    if request.method == 'POST':
        form = IssuingAuthForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Issuing Authority added successfully!")  # Success message
        else:
            messages.error(request, "There was an error adding the Issuing Authority.")  # Error message

    form = IssuingAuthForm()
    issuingauth_form = {
        'form': form,
        'issuingauth_data': IssuingAuth.objects.all(),
    }
    return render(request, 'issuingAuth.html', issuingauth_form)

# Edit Issuing Authority
@login_required
@role_required('admin')
def edit_issuingauth(request, auth_id):
    issuingauth = get_object_or_404(IssuingAuth, auth_id=auth_id)

    if request.method == 'POST':
        form = IssuingAuthForm(request.POST, instance=issuingauth)
        if form.is_valid():
            form.save()
            messages.success(request, "Issuing Authority updated successfully!")  # Success message
            return redirect('issuingauth')
        else:
            messages.error(request, "There was an error updating the Issuing Authority.")  # Error message

    form = IssuingAuthForm(instance=issuingauth)
    return render(request, 'issuingAuth.html', {'form': form, 'issuingauth_data': IssuingAuth.objects.all()})

# Delete Issuing Authority
@login_required
@role_required('admin')
def delete_issuingauth(request, auth_id):
    issuingauth = get_object_or_404(IssuingAuth, auth_id=auth_id)
    if request.method == 'POST':  # Ensure deletion occurs only on POST
        issuingauth.delete()
        messages.success(request, "Issuing Authority deleted successfully!")  # Success message
        return redirect('issuingauth')

    # Optional: If you want to confirm deletion, render a delete confirmation page.
    return render(request, 'delete_confirmation.html', {'object': issuingauth})

#Entity
@login_required
@role_required('admin')
def entity(request):
    if request.method == 'POST':
        form = EntityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Entity added successfully!")
        else:
            messages.error(request, "There was an error adding the entity.")

    form = EntityForm()
    entity_form = {
        'form': form,
        'entity_data': Entity.objects.all(),
    }
    return render(request, 'entity.html', entity_form)

@login_required
@role_required('admin')
def edit_entity(request, entity_id):
    entity = get_object_or_404(Entity, entity_id=entity_id)

    if request.method == 'POST':
        form = EntityForm(request.POST, instance=entity)
        if form.is_valid():
            form.save()
            messages.success(request, "Entity updated successfully!")
            return redirect('entity')
        else:
            messages.error(request, "There was an error updating the entity.")

    form = EntityForm(instance=entity)
    return render(request, 'entity.html', {'form': form, 'entity_data': Entity.objects.all()})

@login_required
@role_required('admin')
def delete_entity(request, entity_id):
    entity = get_object_or_404(Entity, entity_id=entity_id)
    if request.method == 'POST':
        entity.delete()
        messages.success(request, "Entity deleted successfully!")
        return redirect('entity')

@login_required
@role_required('admin')
def licenseperiod(request):
    if request.method == 'POST':
        form = LicenseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "License period added successfully!")
        else:
            messages.error(request, "There was an error adding the license period.")

    form = LicenseForm()
    license_form = {
        'form': form,
        'license_data': LicensePeriod.objects.all(),
    }
    return render(request, 'license.html', license_form)

@login_required
@role_required('admin')
def edit_licenseperiod(request, license_id):
    licenseperiod = get_object_or_404(LicensePeriod, license_id=license_id)

    if request.method == 'POST':
        form = LicenseForm(request.POST, instance=licenseperiod)
        if form.is_valid():
            form.save()
            messages.success(request, "License period updated successfully!")
            return redirect('licenseperiod')
        else:
            messages.error(request, "There was an error updating the license period.")

    form = LicenseForm(instance=licenseperiod)
    return render(request, 'license.html', {'form': form, 'license_data': LicensePeriod.objects.all()})

@login_required
@role_required('admin')
def delete_licenseperiod(request, license_id):
    licenseperiod = get_object_or_404(LicensePeriod, license_id=license_id)
    if request.method == 'POST':
        licenseperiod.delete()
        messages.success(request, "License period deleted successfully!")
        return redirect('licenseperiod')

#Forms
@login_required
@role_required('admin')
def forms(request):
    if request.method == 'POST':
        form = FormsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Form added successfully!")
        else:
            messages.error(request, "There was an error adding the form.")

    form = FormsForm()
    forms_data = Forms.objects.all()
    return render(request, 'forms.html', {'form': form, 'forms_data': forms_data})

@login_required
@role_required('admin')
def edit_forms(request, form_id):
    form_obj = get_object_or_404(Forms, form_id=form_id)

    if request.method == 'POST':
        form = FormsForm(request.POST, instance=form_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Form updated successfully!")
            return redirect('forms')
        else:
            messages.error(request, "There was an error updating the form.")

    form = FormsForm(instance=form_obj)
    return render(request, 'forms.html', {'form': form, 'forms_data': Forms.objects.all()})

@login_required
@role_required('admin')
def delete_forms(request, form_id):
    form_obj = get_object_or_404(Forms, form_id=form_id)
    if request.method == 'POST':
        form_obj.delete()
        messages.success(request, "Form deleted successfully!")
        return redirect('forms')


#Type
@login_required
@role_required('admin')
def type(request):
    if request.method == 'POST':
        form = TypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Type added successfully!")
        else:
            messages.error(request, "There was an error adding the type.")

    form = TypeForm()
    type_form = {
        'form': form,
        'type_data': Type.objects.all(),
    }
    return render(request, 'type.html', type_form)

@login_required
@role_required('admin')
def edit_type(request, type_id):
    type_obj = get_object_or_404(Type, type_id=type_id)

    if request.method == 'POST':
        form = TypeForm(request.POST, instance=type_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Type updated successfully!")
            return redirect('type')
        else:
            messages.error(request, "There was an error updating the type.")

    form = TypeForm(instance=type_obj)
    return render(request, 'type.html', {'form': form, 'type_data': Type.objects.all()})

@login_required
@role_required('admin')
def delete_type(request, type_id):
    type_obj = get_object_or_404(Type, type_id=type_id)
    if request.method == 'POST':
        type_obj.delete()
        messages.success(request, "Type deleted successfully!")
        return redirect('type')



#Shelf number
@login_required
@role_required('admin')
def shelf(request):
    if request.method == 'POST':
        form = ShelfForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Shelf added successfully!")
        else:
            messages.error(request, "There was an error adding the shelf.")

    form = ShelfForm()
    shelf_form = {
        'form': form,
        'shelf_data': Shelf.objects.all(),
    }
    return render(request, 'shelf.html', shelf_form)

@login_required
@role_required('admin')
def edit_shelf(request, shelf_id):
    shelf_obj = get_object_or_404(Shelf, shelf_id=shelf_id)

    if request.method == 'POST':
        form = ShelfForm(request.POST, instance=shelf_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Shelf updated successfully!")
            return redirect('shelf')
        else:
            messages.error(request, "There was an error updating the shelf.")

    form = ShelfForm(instance=shelf_obj)
    return render(request, 'shelf.html', {'form': form, 'shelf_data': Shelf.objects.all()})

@login_required
@role_required('admin')
def delete_shelf(request, shelf_id):
    shelf_obj = get_object_or_404(Shelf, shelf_id=shelf_id)
    if request.method == 'POST':
        shelf_obj.delete()
        messages.success(request, "Shelf deleted successfully!")
        return redirect('shelf')





# #In_out table
# def inout(request):
#     if request.method == 'POST':
#         form = InOutForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "✅ Direction updated successfully!")
#             return redirect('inout')
#         else:
#             # Return form with errors
#             return render(request, 'inout.html', {
#                 'form': form,
#                 'inout_data': InOut.objects.all()
#             })
#     else:
#         form = InOutForm(initial={'direction': 'outgoing'})  #  default "outgoing"
            

#     form = InOutForm()
#     inout_form = {
#         'form' : form,
#         'inout_data' : InOut.objects.all(),
#     }
#     return render(request, 'inout.html', inout_form)




###############################################################################################################################

#newentry



from django.utils.timezone import now


@login_required
@role_required('admin')
def internal_newentry(request):
    dsc_type = "internal"
    form_new = InternalDSCForm()
    form_docs = DocsForm(dsc_type='Internal')
    existing_files = {}
    extra_fields = DSCExtraField.objects.filter(Q(dsc_type="both") | Q(dsc_type=dsc_type)).order_by('id')
  # ✅ Fetch all extra fields for both types
    extra_data = {}  # Store extra field values

    # Auto-generate next DSC number and make it readonly in the view
    # Get last internal DSC with internal type
    try:
        internal_type = get_object_or_404(Type, type_name='Internal')
        last_internal_dsc = DSCMaster.objects.filter(dsc_type=internal_type).order_by('-dsc_id').first()
        
        if last_internal_dsc and last_internal_dsc.dsc_number:
            # Extract numeric part if the format is like "INT-001"
            if '-' in last_internal_dsc.dsc_number:
                prefix, number = last_internal_dsc.dsc_number.split('-')
                new_number = int(number) + 1
                next_dsc_number = f"{prefix}-{new_number:03d}"
            else:
                # If it's just a number, increment it
                next_dsc_number = str(int(last_internal_dsc.dsc_number) + 1)
        else:
            # First entry
            next_dsc_number = "1500"
            
        # Initialize the field with auto-generated value
        form_new.fields['dsc_number'].initial = next_dsc_number
        # Make the field readonly
        form_new.fields['dsc_number'].widget.attrs['readonly'] = True
    except Exception as e:
        # Handle any exceptions
        pass


    if request.method == 'POST':

        # Get the generated DSC number from the form
        generated_dsc_number = request.POST.get('dsc_number')
        form_new = InternalDSCForm(request.POST)
        form_docs = DocsForm(request.POST, request.FILES, dsc_type='Internal')

        
        # ✅ Handle extra fields
    # for field in extra_fields:
    #     extra_data[field.field_name] = request.POST.get(field.field_name, '')


        # Make sure the field is readonly in POST too
        form_new.fields['dsc_number'].widget.attrs['readonly'] = True

        if form_new.is_valid() and form_docs.is_valid():
            form_new_instance = form_new.save(commit=False)

            # Ensure we use the generated DSC number
            form_new_instance.dsc_number = generated_dsc_number

            # ✅ Set the logged-in user as user_id
            form_new_instance.user_id = request.user  # Assign the logged-in user directly

            # ✅ Automatically set DSC Type to Internal
            internal_type = get_object_or_404(Type, type_name='Internal')  
            form_new_instance.dsc_type = internal_type

            # ✅ Handle expiry status
            issued_date = form_new.cleaned_data['issued_date']
            license_period = form_new.cleaned_data['license_period'].no_of_years
            expiry_date = issued_date.replace(year=issued_date.year + license_period)

            form_new_instance.status = "Expired" if expiry_date < now().date() else "Active"
            form_new_instance.save()

            # Initialize extra_fields as empty dict if None
            if not hasattr(form_new_instance, 'extra_fields'):
                form_new_instance.extra_fields = {}
            
            # Save extra fields to both DSCExtraData AND DSCMaster.extra_fields
            for field in extra_fields:
                field_value = request.POST.get(field.field_name)
                if field_value:
                    # Save to DSCExtraData
                    DSCExtraData.objects.update_or_create(
                        dsc_number=form_new_instance,
                        field_name=field.field_name,
                        defaults={'value': field_value}
                    )
                    
                    # Also save to DSCMaster.extra_fields JSON field
                    if field.field_type == "boolean":
                        field_value = True if field_value == "on" else False
                    form_new_instance.extra_fields[field.field_name] = field_value

            # Save the instance after updating extra_fields
            form_new_instance.save()                    
            
            
            # ✅ Save uploaded documents
            form_docs_instance = form_docs.save(commit=False)
            form_docs_instance.dsc_number = form_new_instance
            form_docs_instance.save()

            # ✅ Automatically Create InOut Entry
            InOut.objects.create(
                dsc_number=form_new_instance,
                direction="in",
                initiated=False,
                action_completed=False
            )

            messages.success(request, "✅ Internal DSC entry added successfully!")
            return redirect('map_dsc_entity')

        else:
            messages.error(request, "❌ Please fill all required fields for Internal DSC.")
            print("Form Errors (form_new):", form_new.errors)
            print("Form Errors (form_docs):", form_docs.errors)


            # ✅ Store uploaded file names to keep them after a validation failure
            for field in ['aadhar_path', 'pan_path', 'pp_path']:
                if field in request.FILES:
                    existing_files[field] = request.FILES[field].name  # ✅ Store filename

    return render(request, 'internal_newentry.html', {
        'form_new': form_new, 
        'form_docs': form_docs,
        'existing_files': existing_files,  # ✅ Pass retained file names to template
        'extra_fields': extra_fields,  # ✅ Pass extra fields to template
    })





@login_required
@role_required('admin')
def external_newentry(request):
    dsc_type = "external"
    form_new = ExternalDSCForm()
    form_docs = DocsForm(dsc_type='External')
    extra_fields = DSCExtraField.objects.filter(Q(dsc_type="both") | Q(dsc_type=dsc_type)).order_by('id')
  # ✅ Fetch all extra fields for both types
    extra_data = {}  # Store extra field values

    
    if request.method == 'POST':
        form_new = ExternalDSCForm(request.POST)
        form_docs = DocsForm(request.POST, request.FILES, dsc_type='External')  # ✅ Pass dsc_type


        if form_new.is_valid() and form_docs.is_valid():
            form_new_instance = form_new.save(commit=False)
            
            # ✅ Set the logged-in user as user_id
            form_new_instance.user_id = request.user  

            # ✅ Automatically set DSC Type to External
            external_type = get_object_or_404(Type, type_name='External')
            form_new_instance.dsc_type = external_type

            # ✅ Handle optional fields correctly
            form_new_instance.dsc_class = form_new.cleaned_data.get('dsc_class') or None
            form_new_instance.issuing_auth = form_new.cleaned_data.get('issuing_auth') or None
            form_new_instance.type = form_new.cleaned_data.get('type') or None

            # ✅ Handle expiry status
            issued_date = form_new.cleaned_data['issued_date']
            license_period = form_new.cleaned_data['license_period'].no_of_years
            expiry_date = issued_date.replace(year=issued_date.year + license_period)

            form_new_instance.status = "Expired" if expiry_date < now().date() else "Active"
            form_new_instance.save()

            # Initialize extra_fields if None
            if not hasattr(form_new_instance, 'extra_fields'):
                form_new_instance.extra_fields = {}
            
            # Save extra fields to both storage locations
            for field in extra_fields:
                field_value = request.POST.get(field.field_name)
                if field_value:
                    # Save to DSCExtraData
                    DSCExtraData.objects.update_or_create(
                        dsc_number=form_new_instance,
                        field_name=field.field_name,
                        defaults={'value': field_value}
                    )
                    
                    # Also save to DSCMaster.extra_fields JSON field
                    if field.field_type == "boolean":
                        field_value = True if field_value == "on" else False
                    form_new_instance.extra_fields[field.field_name] = field_value
            
            form_new_instance.save()
                     
            # Save documents                     
            form_docs_instance = form_docs.save(commit=False)
            form_docs_instance.dsc_number = form_new_instance
            form_docs_instance.save()

            messages.success(request, "✅ External DSC entry added successfully!")
            return redirect('map_dsc_entity')

        else:
            print("❌ External DSC Form Errors:", form_new.errors)
            print("❌ Documents Form Errors:", form_docs.errors)
            messages.error(request, "❌ Please fill in all required fields for External DSC. Please check the form. There are some errors.")

    else:
        form_new = ExternalDSCForm()
        form_docs = DocsForm(dsc_type='external')  # ✅ Pass dsc_type

    return render(request, 'external_newentry.html', {'form_new': form_new, 'form_docs': form_docs,'extra_fields': extra_fields})# ✅ Pass extra fields to template



@login_required
@role_required('admin')
def map_dsc_entity(request):
    """Handles mapping of DSCs to Entities and displays mapped records."""
    dsc_numbers = DSCMaster.objects.all().order_by('dsc_number')

    if request.method == 'POST':
        form = DSCEntityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Entity mapped to DSC successfully!")
            return redirect('map_dsc_entity')
        else:
            messages.error(request, "❌ Error in mapping DSC to Entity.")
    else:
        form = DSCEntityForm()

    selected_entity_id = request.GET.get("entity_id")  # Get selected entity from dropdown
    entities = Entity.objects.all()  # Fetch all available entities

    # ✅ Fetch only MAPPED DSC-Entities
    mappings = DSCEntity.objects.select_related("dsc_number", "entity")

    # ✅ Apply entity filter if selected
    if selected_entity_id:
        mappings = mappings.filter(entity_id=selected_entity_id)

    return render(request, "map_dsc_entity.html", {
        "form": form,
        "entities": entities,
        "mappings": mappings,
        "selected_entity_id": selected_entity_id,
        "dsc_numbers": dsc_numbers,
    })


@login_required
@role_required('admin')
def delete_mapping(request, mapping_id):
    """Deletes a DSC-Entity mapping."""
    mapping = get_object_or_404(DSCEntity, id=mapping_id)
    mapping.delete()
    messages.success(request, "✅ Mapping deleted successfully!")
    return redirect('map_dsc_entity')






from django.utils.timezone import now, timedelta

from django.utils.timezone import now
from datetime import timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from home.models import DSCMaster, DSCRenewal, Docs, InOut
from home.forms import DSCRenewalForm



@login_required
def api_dsc_list(request):
    dscs = DSCMaster.objects.all().order_by('dsc_number').values(
        'dsc_number', 'full_name', 'pan_no', 'email_id', 'phone_no', 
        'expiry_date', 'dsc_type__type_name'
    )
    
    # Convert to list and rename fields for consistency
    dsc_list = []
    for dsc in dscs:
        dsc_list.append({
            'dsc_number': dsc['dsc_number'],
            'full_name': dsc['full_name'],
            'pan_no': dsc['pan_no'],
            'email_id': dsc['email_id'],
            'phone_no': dsc['phone_no'],
            'expiry_date': dsc['expiry_date'].strftime('%d/%m/%Y') if dsc['expiry_date'] else None,
            'dsc_type': dsc['dsc_type__type_name']
        })
    
    return JsonResponse(dsc_list, safe=False)

# views.py

@login_required
@role_required('admin')
def renewal(request):
    search_query = request.GET.get('search', '')
    dsc_number = request.GET.get('dsc_number')
    
    # Start with all DSCs ordered by DSC number
    dsc_numbers = DSCMaster.objects.all().order_by('dsc_number')
    
    # Apply search filter if provided
    if search_query:
        dsc_numbers = dsc_numbers.filter(
            Q(dsc_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(pan_no__icontains=search_query) |
            Q(email_id__icontains=search_query) |
            Q(phone_no__icontains=search_query)
        )
    
    dsc_master = None
    dsc_form = None
    doc_form = None
    documents = None
    extra_fields_with_values = []

    if request.method == 'POST':
        dsc_number = request.POST.get('dsc_number')
        if not dsc_number:
            messages.error(request, "❌ Please select a DSC number.")
            return redirect('renewal')

        dsc_master = get_object_or_404(DSCMaster, dsc_number=dsc_number)
        old_password = dsc_master.password  

        # Capture data before renewal for history
        previous_data = {
            'dsc_number': dsc_master.dsc_number,
            'full_name': dsc_master.full_name,
            'issued_date': dsc_master.issued_date.isoformat(),
            'expiry_date': dsc_master.expiry_date.isoformat() if dsc_master.expiry_date else None,
            'license_period': dsc_master.license_period.no_of_years if dsc_master.license_period else None,
            'password': dsc_master.password,
            'pan_no': dsc_master.pan_no,
            'dsc_class': dsc_master.dsc_class.class_name if dsc_master.dsc_class else None,
            'email_id': dsc_master.email_id,
            'phone_no': dsc_master.phone_no,
            'issuing_auth': dsc_master.issuing_auth.auth_name if dsc_master.issuing_auth else None,
            'ref_name': dsc_master.ref_name,
            'ref_contact': dsc_master.ref_contact,
            'type': dsc_master.type,
            'remarks': dsc_master.remarks,
            'extra_fields': dsc_master.extra_fields,
        }
        
        # Capture document info before renewal
        current_docs = Docs.objects.filter(dsc_number=dsc_master).first()
        previous_docs = {
            'aadhar': current_docs.aadhar_path.url if current_docs and current_docs.aadhar_path else None,
            'pan': current_docs.pan_path.url if current_docs and current_docs.pan_path else None,
            'pp': current_docs.pp_path.url if current_docs and current_docs.pp_path else None,
        }

        # Get extra fields for this DSC type
        extra_fields = DSCExtraField.objects.filter(
            Q(dsc_type=dsc_master.dsc_type.type_name.lower()) | Q(dsc_type="both"))

        # Choose the correct form based on DSC type
        if dsc_master.dsc_type.type_name.lower() == "internal":
            DscFormClass = InternalDSCForm
        else:
            DscFormClass = ExternalDSCForm  

        dsc_form = DscFormClass(request.POST, instance=dsc_master)
        documents = Docs.objects.filter(dsc_number=dsc_master)
        doc_form = DocsForm(request.POST, request.FILES)

        if dsc_form.is_valid() and doc_form.is_valid():
            # Process extra fields - update both storage locations
            extra_data = {}
            for field in extra_fields:
                field_value = request.POST.get(field.field_name)
                if field_value:
                    if field.field_type == "boolean":
                        field_value = True if field_value == "on" else False
                    
                    # Update JSON field
                    extra_data[field.field_name] = field_value
                    
                    # Update or create DSCExtraData record
                    DSCExtraData.objects.update_or_create(
                        dsc_number=dsc_master,
                        field_name=field.field_name,
                        defaults={'value': str(field_value)}
                    )
            
            dsc_master.extra_fields = extra_data

            # Handle password
            new_password = dsc_form.cleaned_data.get('password')
            if not new_password:
                dsc_master.password = old_password

            # Calculate new expiry date
            new_issued_date = dsc_form.cleaned_data.get('issued_date')
            new_license_period = dsc_form.cleaned_data.get('license_period')
            new_expiry_date = new_issued_date + timedelta(days=new_license_period.no_of_years * 365)
            dsc_master.expiry_date = new_expiry_date
            dsc_master.status = "Active" if new_expiry_date > now().date() else "Expired"
            dsc_master.save()

            # Handle document updates
            new_docs = {}
            if current_docs:
                if 'aadhar_path' in request.FILES:
                    current_docs.aadhar_path = request.FILES['aadhar_path']
                if 'pan_path' in request.FILES:
                    current_docs.pan_path = request.FILES['pan_path']
                if 'pp_path' in request.FILES:
                    current_docs.pp_path = request.FILES['pp_path']
                current_docs.save()
                
                new_docs = {
                    'aadhar': current_docs.aadhar_path.url if current_docs.aadhar_path else None,
                    'pan': current_docs.pan_path.url if current_docs.pan_path else None,
                    'pp': current_docs.pp_path.url if current_docs.pp_path else None,
                }
            else:
                doc_instance = doc_form.save(commit=False)
                doc_instance.dsc_number = dsc_master
                doc_instance.save()
                new_docs = {
                    'aadhar': doc_instance.aadhar_path.url if doc_instance.aadhar_path else None,
                    'pan': doc_instance.pan_path.url if doc_instance.pan_path else None,
                    'pp': doc_instance.pp_path.url if doc_instance.pp_path else None,
                }

            # Prepare new data for history
            new_data = {
                'dsc_number': dsc_form.cleaned_data.get('dsc_number'),
                'full_name': dsc_form.cleaned_data.get('full_name'),
                'issued_date': new_issued_date.isoformat(),
                'expiry_date': new_expiry_date.isoformat(),
                'license_period': new_license_period.no_of_years,
                'password': new_password if new_password else old_password,
                'pan_no': dsc_form.cleaned_data.get('pan_no'),
                'dsc_class': dsc_form.cleaned_data.get('dsc_class').class_name if dsc_form.cleaned_data.get('dsc_class') else None,
                'email_id': dsc_form.cleaned_data.get('email_id'),
                'phone_no': dsc_form.cleaned_data.get('phone_no'),
                'issuing_auth': dsc_form.cleaned_data.get('issuing_auth').auth_name if dsc_form.cleaned_data.get('issuing_auth') else None,
                'ref_name': dsc_form.cleaned_data.get('ref_name'),
                'ref_contact': dsc_form.cleaned_data.get('ref_contact'),
                'type': dsc_form.cleaned_data.get('type'),
                'remarks': dsc_form.cleaned_data.get('remarks'),
                'extra_fields': extra_data,
            }

            # Create renewal history record
            DSCRenewalHistory.objects.create(
                dsc=dsc_master,
                renewed_by=request.user,
                previous_data=previous_data,
                new_data=new_data,
                previous_documents=previous_docs,
                new_documents=new_docs
            )

            # Remove DSC from "Not Renewing" category if mistakenly added
            DSCRenewal.objects.filter(dsc_number=dsc_master, remarks__isnull=False).delete()

            messages.success(request, "✅ DSC renewed successfully!")
            return redirect('renewal')
        else:
            messages.error(request, "❌ Please fill in all required fields.")
            print("Form Errors (dsc_form):", dsc_form.errors)
            print("Form Errors (doc_form):", doc_form.errors)

    # Handle GET request
    dsc_number = request.GET.get('dsc_number')
    if dsc_number:
        dsc_master = get_object_or_404(DSCMaster, dsc_number=dsc_number)
        
        # Get extra fields for this DSC type
        extra_fields = DSCExtraField.objects.filter(
            Q(dsc_type=dsc_master.dsc_type.type_name.lower()) | Q(dsc_type="both"))
        
        # Prepare extra fields with values - check both sources
        extra_fields_with_values = []
        for field in extra_fields:
            # First check JSON field
            field_value = dsc_master.extra_fields.get(field.field_name) if dsc_master.extra_fields else None
            
            # If not found in JSON, check DSCExtraData
            if field_value is None:
                try:
                    extra_data = DSCExtraData.objects.get(
                        dsc_number=dsc_master,
                        field_name=field.field_name
                    )
                    field_value = extra_data.value
                    # Update JSON field for consistency
                    if not dsc_master.extra_fields:
                        dsc_master.extra_fields = {}
                    dsc_master.extra_fields[field.field_name] = field_value
                    dsc_master.save()
                except DSCExtraData.DoesNotExist:
                    field_value = ''
            
            extra_fields_with_values.append({
                'field_name': field.field_name,
                'display_name': field.label or field.field_name,
                'value': field_value,
                'field_type': field.field_type,
                'is_required': field.required,
                'choices': field.get_FIELD_TYPE_CHOICES_display() if field.field_type == 'select' else []
            })
        
        # Choose the correct form dynamically
        if dsc_master.dsc_type.type_name.lower() == "internal":
            DscFormClass = InternalDSCForm
        else:
            DscFormClass = ExternalDSCForm
            
        dsc_form = DscFormClass(instance=dsc_master) 
        dsc_form.fields['password'].widget.attrs['value'] = dsc_master.password  # Pre-fill password
        documents = Docs.objects.filter(dsc_number=dsc_master)
        doc_form = DocsForm()  # Form for uploading new documents
        
    return render(request, 'renewal.html', {
        'dsc_master': dsc_master,
        'dsc_numbers': dsc_numbers,
        'dsc_form': dsc_form,
        'doc_form': doc_form,
        'documents': documents,
        'extra_fields_with_values': extra_fields_with_values,
        'history_count': DSCRenewalHistory.objects.filter(dsc=dsc_master).count() if dsc_master else 0,
    })



from datetime import date as datetime_date  # Add this import at the top

@login_required
@role_required('admin')
def additional_details(request):
    """Manage Additional Details for DSCs"""
    dsc_numbers = DSCMaster.objects.all()
    additional_form = None
    dsc_master = None
    previous_additional_details = None
    extra_fields_data = []

    if request.method == 'POST':
        dsc_number = request.POST.get('dsc_number') or request.POST.get('dsc_number_hidden')

        if not dsc_number:
            messages.error(request, "❌ No DSC selected. Please choose a DSC.")
            return redirect('additional_details')

        try:
            dsc_master = DSCMaster.objects.get(dsc_number=dsc_number)
        except DSCMaster.DoesNotExist:
            messages.error(request, f"❌ No DSC found with number {dsc_number}. Please try again.")
            return redirect('additional_details')

        previous_additional_details = DSCRenewal.objects.filter(dsc_number=dsc_master).order_by('-additional_id')
        latest_additional_instance = previous_additional_details.first()

        additional_form = DSCRenewalForm(request.POST, instance=latest_additional_instance)

        if additional_form.is_valid():
            additional_instance = additional_form.save(commit=False)
            additional_instance.dsc_number = dsc_master

            # Save extra fields to JSON field with proper serialization
            additional_instance.extra_fields = {}
            for field in RenewalExtraField.objects.all():
                field_name = field.field_name
                if field_name in additional_form.cleaned_data:
                    value = additional_form.cleaned_data[field_name]
                    
                    # Handle different field types
                    if field.field_type == "boolean":
                        value = bool(value)
                    elif field.field_type == "date" and value:
                        if isinstance(value, datetime_date):
                            value = value.isoformat()  # Convert date to string
                    elif value is None:
                        continue
                        
                    additional_instance.extra_fields[field_name] = value

            additional_instance.save()
            messages.success(request, "✅ Additional details updated successfully!")
            return redirect('additional_details')

    dsc_number = request.GET.get('dsc_number')
    if dsc_number:
        try:
            dsc_master = DSCMaster.objects.get(dsc_number=dsc_number)
            previous_additional_details = DSCRenewal.objects.filter(dsc_number=dsc_master).order_by('-additional_id')
            latest_additional_instance = previous_additional_details.first()

            if latest_additional_instance:
                additional_form = DSCRenewalForm(instance=latest_additional_instance)
                # Deserialize dates from JSON
                if latest_additional_instance.extra_fields:
                    for field in RenewalExtraField.objects.filter(field_type='date'):
                        if field.field_name in latest_additional_instance.extra_fields:
                            date_str = latest_additional_instance.extra_fields[field.field_name]
                            if date_str:
                                try:
                                    # Fixed: Use datetime_date.fromisoformat()
                                    additional_form.initial[field.field_name] = datetime_date.fromisoformat(date_str)
                                except (ValueError, TypeError):
                                    pass
            else:
                additional_form = DSCRenewalForm()

            # Prepare extra fields data for template
            if additional_form:
                for field in RenewalExtraField.objects.all():
                    try:
                        extra_fields_data.append({
                            'name': field.field_name,
                            'field': additional_form[field.field_name],
                            'label': field.label or field.field_name.replace('_', ' ').title(),
                            'required': field.required,
                            'help_text': field.help_text,
                            'field_type': field.field_type  # Add field type to context
                        })
                    except KeyError:
                        continue

        except DSCMaster.DoesNotExist:
            messages.error(request, f"❌ No DSC found with number {dsc_number}. Please try again.")
            return redirect('additional_details')

    return render(request, 'additional_details.html', {
        'dsc_numbers': dsc_numbers,
        'additional_form': additional_form,
        'selected_dsc': dsc_master,
        'previous_additional_details': previous_additional_details,
        'extra_fields_data': extra_fields_data,
    })


class RenewalFieldListView(ListView):
    model = RenewalExtraField
    template_name = 'manage_renewal_fields.html'
    context_object_name = 'extra_fields'
    paginate_by = 10

class RenewalFieldCreateView(CreateView):
    model = RenewalExtraField
    template_name = 'renewal_field_form.html'
    fields = ['field_name', 'field_type', 'label', 'required', 'help_text']
    success_url = reverse_lazy('manage_renewal_fields')

    def form_valid(self, form):
        messages.success(self.request, "Field added successfully!")
        return super().form_valid(form)

class RenewalFieldUpdateView(UpdateView):
    model = RenewalExtraField
    template_name = 'renewal_field_form.html'
    fields = ['field_name', 'field_type', 'label', 'required', 'help_text']
    success_url = reverse_lazy('manage_renewal_fields')

    def form_valid(self, form):
        messages.success(self.request, "Field updated successfully!")
        return super().form_valid(form)

class RenewalFieldDeleteView(DeleteView):
    model = RenewalExtraField
    template_name = 'renewal_field_confirm_delete.html'
    success_url = reverse_lazy('manage_renewal_fields')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Field deleted successfully!")
        return super().delete(request, *args, **kwargs)


# Usage Logs List and Form Submission
# Usage Logs List and Form Submission
@login_required
@role_required('admin')
def usage_logs(request):
    if request.method == "POST":
        print("Received POST data:", request.POST)  # ✅ Debugging

        # ✅ Get POST data
        form_other = request.POST.get("form_other", "").strip()
        platform_other = request.POST.get("platform_other", "").strip()
        form_id = request.POST.get("form_id")
        platform_id = request.POST.get("platform_id")

        # ✅ If "Other" is selected, create a new Form object
        if form_id == "other":
            if not form_other:
                messages.error(request, "❌ Please enter a form name.")
                return redirect("usage_logs")
            form_obj, created = Forms.objects.get_or_create(form_name=form_other)
            form_id = form_obj.pk  # ✅ Use the new Form ID

        # ✅ If "Other" is selected, create a new Platform object
        if platform_id == "other":
            if not platform_other:
                messages.error(request, "❌ Please enter a platform name.")
                return redirect("usage_logs")
            platform_obj, created = Platform.objects.get_or_create(platform_name=platform_other)
            platform_id = platform_obj.pk  # ✅ Use the new Platform ID

        # ✅ Modify request.POST to include new IDs
        mutable_post = request.POST.copy()
        mutable_post["form_id"] = form_id
        mutable_post["platform_id"] = platform_id

        form = UsageLogsForm(mutable_post, request=request)

        if form.is_valid():
            print("✅ Form is valid")  # ✅ Debugging
            usage_log = form.save(commit=False)

            # ✅ Assign current user
            usage_log.requested_by = request.user
            usage_log.save()

            messages.success(request, "✅ Usage log submitted successfully! Waiting for approval.")
            return redirect("usage_logs")

        else:
            print("❌ Form errors:", form.errors)  # ✅ Debugging
            messages.error(request, "❌ Error in submission. Please check the form.")

    else:
        form = UsageLogsForm(request=request)

    usage_logs_data = UsageLogs.objects.all()
        # Filtering logic
    usage_logs_data = UsageLogs.objects.all().select_related(
        'dsc_id', 'form_id', 'platform_id', 'requested_by', 'approved_by_user'
    ).order_by('-date_of_usage')

    search_query = request.GET.get('search')
    if search_query:
        usage_logs_data = usage_logs_data.filter(
            Q(dsc_id__dsc_number__icontains=search_query) |
            Q(form_id__form_name__icontains=search_query) |
            Q(platform_id__platform_name__icontains=search_query) |
            Q(requested_by__username__icontains=search_query)
        )

    status_filter = request.GET.get('status')
    if status_filter:
        usage_logs_data = usage_logs_data.filter(approval_status=status_filter)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        usage_logs_data = usage_logs_data.filter(
            date_of_usage__gte=start_date,
            date_of_usage__lte=end_date
        )


    paginator = Paginator(usage_logs_data, 15)  
    page_number = request.GET.get('page')
    usage_logs_data = paginator.get_page(page_number)

    return render(request, "usagelogs.html", {
        "form": form,
        "usage_logs_data": usage_logs_data,
    })




# Delete Usage Log
@login_required
@role_required('admin')
def delete_usage_log(request, log_id):
    usage_log = get_object_or_404(UsageLogs, log_id=log_id)

    usage_log.delete()
    messages.success(request, "Usage log deleted successfully!")
    return redirect('usage_logs')







@login_required
@role_required('admin')
def search_dsc(request):
    query = request.GET.get('q', '').strip()  # Get the search query
    search_type = request.GET.get('type', 'name')  # Determine search type

    results = []    
    if query:
        if search_type == 'name':
            results = DSCMaster.objects.filter(full_name__icontains=query)
        elif search_type == 'entity':
            results = DSCMaster.objects.filter(dscentity_set__entity__entity_name__icontains=query).distinct()
        elif search_type == 'pan':
            results = DSCMaster.objects.filter(pan_no__icontains=query)
        elif search_type == 'expiry_date':
            results = DSCMaster.objects.filter(expiry_date__icontains=query)
        elif search_type == 'status':
            results = DSCMaster.objects.filter(status__icontains=query)  # Active / Expired

    # ✅ Attach latest **completed** direction to each DSC result
    for dsc in results:
        latest_completed_inout = (
            InOut.objects.filter(dsc_number=dsc, action_completed=True)
            .order_by('-in_out_id')
            .first()
        )
        dsc.latest_direction = latest_completed_inout.direction if latest_completed_inout else "in"

    return render(request, 'search_results.html', {
        'query': query,
        'search_type': search_type,
        'results': results
    })


@login_required
@role_required('admin')
def dsc_display(request, dsc_number):
    dsc_master = get_object_or_404(DSCMaster, dsc_number=dsc_number)

    # ✅ Get the latest completed transaction (Actual last known direction)
    latest_completed = (
        InOut.objects.filter(dsc_number=dsc_master, action_completed=True)
        .order_by('-in_out_id')
        .first()
    )

    # ✅ Get the latest initiated transaction (Ongoing initiation)
    latest_initiated = (
        InOut.objects.filter(dsc_number=dsc_master, initiated=True, action_completed=False)
        .order_by('-in_out_id')
        .first()
    )

    # ✅ Correct Current Direction Logic
    if latest_completed:
        current_direction = latest_completed.direction  # ✅ Use last **completed** direction
    else:
        current_direction = "in"  # ✅ Default for new DSCs

    # ✅ Fetch related data
    documents = Docs.objects.filter(dsc_number=dsc_master)
    entities = (
        DSCEntity.objects.filter(dsc_number=dsc_master)
        .values_list('entity__entity_name', flat=True)
    )

    # Process additional fields - check both JSON field and DSCExtraData
    additional_fields = []
    extra_fields = DSCExtraField.objects.filter(
        Q(dsc_type=dsc_master.dsc_type.type_name.lower()) | Q(dsc_type="both")
    )
    
    for field in extra_fields:
        # First check JSON field
        value = dsc_master.extra_fields.get(field.field_name) if dsc_master.extra_fields else None
        
        # If not found in JSON, check DSCExtraData
        if value is None:
            try:
                extra_data = DSCExtraData.objects.get(
                    dsc_number=dsc_master,
                    field_name=field.field_name
                )
                value = extra_data.value
                # Update JSON field for consistency
                if not dsc_master.extra_fields:
                    dsc_master.extra_fields = {}
                dsc_master.extra_fields[field.field_name] = value
                dsc_master.save()
            except DSCExtraData.DoesNotExist:
                value = None
        
        additional_fields.append({
            'name': field.label or field.field_name.replace('_', ' ').title(),
            'value': value if value not in [None, ""] else "Not provided"
        })
    
    print("\nDEBUG - Fields to Display:", additional_fields)  # Verify before rendering
    
    return render(request, 'dscDisplay.html', {
        'dsc_master': dsc_master,
        'entities': entities,
        'documents': documents,
        'current_direction': current_direction,  # ✅ Pass to template
        'additional_fields': additional_fields,
    })







User = get_user_model()  # Get CustomUser model

@login_required
@role_required('admin')
def manage_users(request):
    if request.user.role != 'admin':  
        messages.error(request, "You do not have permission to access this page.")
        return redirect('admin_dashboard')  

    users = CustomUser.objects.all()
    return render(request, 'manage_users.html', {'users': users})

@login_required
@role_required('admin')
def add_user(request):
    if request.user.role != 'admin':
        messages.error(request, "You do not have permission to add users.")
        return redirect('manage_users')

    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ User added successfully!")
            return redirect('manage_users')  # ✅ Redirect to Manage Users
    else:
        form = UserForm()

    return render(request, 'add_user.html', {'form': form})

@login_required
@role_required('admin')
def edit_user(request, user_id):
    if request.user.role != 'admin':
        messages.error(request, "You do not have permission to edit users.")
        return redirect('manage_users')

    user = get_object_or_404(User, id=user_id)
    if request.method == "POST":
        form = EditUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ User updated successfully!")
            return redirect('manage_users')  # ✅ Redirect to Manage Users
    else:
        form = EditUserForm(instance=user)

    return render(request, 'edit_user.html', {'form': form, 'user': user})

@login_required
@role_required('admin')
def delete_user(request, user_id):
    if request.user.role != 'admin':
        messages.error(request, "You do not have permission to delete users.")
        return redirect('manage_users')

    user = get_object_or_404(User, id=user_id)
    if request.method == "POST":
        user.delete()
        messages.success(request, "✅ User deleted successfully!")
        return redirect('manage_users')  # ✅ Redirect to Manage Users

    return render(request, 'delete_user.html', {'user': user})






def login_view(request):
    if request.method == "POST":
        form = LoginForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect_based_on_role(user)  # Redirect based on role
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})



@login_required
def logout_view(request):
    logout(request)
    return redirect('login')







def redirect_based_on_role(user):
    if user.is_superuser:  # 🔹 Explicit check for Django superuser
        return redirect('admin_dashboard')
    elif user.role == 'admin':
        return redirect('admin_dashboard')
    elif user.role == 'general-user':
        return redirect('general_user_dashboard')
    elif user.role == 'approver':
        return redirect('approver_dashboard')
    return redirect('login')  # Default redirect if role is invalid







@login_required
@role_required('admin')
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')

@login_required
@role_required('general-user')
def general_user_dashboard(request):
    return render(request, 'general_user/general_user_dashboard.html')

@login_required
@role_required('approver')
def approver_dashboard(request):
    return render(request, 'approver/approver_dashboard.html')


from django.db.models import Prefetch
@login_required
@role_required('admin')
def admin_dashboard(request):
    today = date.today()

    # ✅ Fetch related entities in a single query
    dsc_entities = DSCEntity.objects.select_related("entity")
    
    # ✅ Fetch DSCs that are still "Expired" and "Pending Renewal"
    expired_pending_renewal = DSCMaster.objects.prefetch_related(
    Prefetch("dscentity_set", queryset=dsc_entities)
    ).filter(
        expiry_date__lt=today, 
        status="Expired"
    ).exclude(
        dsc_id__in=DSCRenewal.objects.filter(
            remarks__isnull=False  # ✅ Allow any remarks, not just "not renewing"
        ).values_list('dsc_number__dsc_id', flat=True)  
    )

    # ✅ Expired DSCs Marked as "Not Renewing"
    expired_not_renewing = DSCRenewal.objects.select_related("dsc_number").filter(
        remarks__isnull=False,  # ✅ Allow any remarks instead of just "not renewing"
        dsc_number__status="Expired"
    ).prefetch_related("dsc_number__dscentity_set__entity")


    # ✅ DSCs Expiring in Next 30 Days
    upcoming_expiring = DSCMaster.objects.filter(
        expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30), status="Active"
    )

    # ✅ DSCs Missing Additional Details (No Entry in DSCRenewal)
    missing_details = DSCMaster.objects.filter(
    Q(renewals__gst_reg_date__isnull=True) | 
    Q(renewals__mca_reg_date__isnull=True) | 
    Q(renewals__it_reg_date__isnull=True)
).distinct()

    context = {
        "expired_pending_renewal": expired_pending_renewal,  # ✅ FIXED
        "expired_not_renewing": expired_not_renewing,  
        "upcoming_expiring": upcoming_expiring,
        "missing_details": missing_details,
    }
    return render(request, "admin_dashboard.html", context)



@login_required
@role_required('admin')
def mark_not_renewing(request, dsc_number):
    """ Move a DSC from 'Expired Pending Renewal' to 'Expired Not Renewing' """
    dsc = get_object_or_404(DSCMaster, dsc_number=dsc_number)

    if request.method == "POST":
        remarks = request.POST.get("remarks", "").strip()

        if not remarks:
            messages.error(request, "❌ Remarks are required!")
            return redirect("admin_dashboard")

        # ✅ Delete duplicate entries if they exist
        DSCRenewal.objects.filter(dsc_number=dsc).delete()

        # ✅ Create a new entry in DSCRenewal (marks it as Not Renewing)
        DSCRenewal.objects.create(dsc_number=dsc, remarks=remarks)

        messages.success(request, f"✅ DSC {dsc_number} moved to 'Not Renewing'.")
    
    return redirect("admin_dashboard")




@login_required
@role_required('admin')
def mark_pending_renewal(request, dsc_number):
    """ Move a DSC from 'Expired Not Renewing' back to 'Expired Pending Renewal' """
    renewal_entry = get_object_or_404(DSCRenewal, dsc_number__dsc_number=dsc_number)

    if request.method == "POST":
        # ✅ Delete the entry in DSCRenewal (removes it from 'Not Renewing')
        renewal_entry.delete()

        messages.success(request, f"✅ DSC {dsc_number} moved back to 'Pending Renewal'.")
    
    return redirect("admin_dashboard")








@login_required
@role_required('admin')
def initiate_dsc(request):
    """Handles initiation of DSC movement."""
    if request.method == 'POST':
        form = InitiationForm(request.POST)

        if form.is_valid():
            initiation = form.save(commit=False)
            initiation.initiated_by = request.user
            initiation.initiated = True  # ✅ Mark as initiated

            # ✅ Auto-fill requester_name if "Self" was selected
            if request.POST.get('requester_type_value') == "self":
                initiation.requester_name = request.user.username  # Assign logged-in user's name

            if initiation.initiation_date is None:
                initiation.initiation_date = now()

            initiation.save()
            messages.success(request, f"✅ DSC {initiation.dsc_number} has been initiated.")
            return redirect('initiate_dsc')
        else:
            messages.error(request, "❌ Form submission failed. Please check the fields.")

    else:
        form = InitiationForm()

    # ✅ Fetch **only pending** initiated logs (not completed ones)
    initiated_logs = InOut.objects.filter(initiated=True, action_completed=False, initiated_by__isnull=False).select_related('dsc_number')


    # ✅ Fetch last completed direction for each DSC
    for log in initiated_logs:
        last_completed = (
            InOut.objects.filter(dsc_number=log.dsc_number, action_completed=True)
            .order_by('-action_date', '-action_time')
            .first()
        )
        log.last_completed_direction = last_completed.direction if last_completed else None

    dsc_list = DSCMaster.objects.all()

    return render(request, 'initiation.html', {
        'form': form,
        'initiated_logs': initiated_logs,
        'dsc_list': dsc_list,
    })



@login_required
@role_required('admin')
def delivery_collection(request):
    """Handles marking of DSC as Delivered or Collected."""
    if request.method == 'POST':
        log_id = request.POST.get('log_id')

        if not log_id:
            messages.error(request, "❌ No DSC was selected.")
            return redirect('delivery_collection')

        log = get_object_or_404(InOut, in_out_id=log_id)
        form = DeliveryCollectionForm(request.POST, instance=log)

        if form.is_valid():
            delivery = form.save(commit=False)
            delivery.action_completed = True  # ✅ Mark as completed

            if not delivery.action_date:
                delivery.action_date = now().date()

            if not delivery.action_time:
                delivery.action_time = now().time()

            # ✅ Only change direction after successful delivery/collection
            if log.direction == "in":  
                delivery.direction = "out"
            else:
                delivery.direction = "in"

            delivery.save()
            messages.success(request, f"✅ DSC {delivery.dsc_number} has been successfully delivered/collected.")
            return redirect('delivery_collection')
        else:
            messages.error(request, "❌ There was an error in form submission.")

    else:
        form = DeliveryCollectionForm()

    
    # ✅ Fetch pending logs and ensure each log has its initiated direction
    pending_logs = InOut.objects.filter(initiated=True, action_completed=False, initiated_by__isnull=False).annotate(
        initiated_direction=F('direction')  # ✅ Get the direction at the time of initiation
    )

    # ✅ Add filtering functionality
    direction_filter = request.GET.get('direction')
    dsc_search = request.GET.get('dsc_search')

    if direction_filter:
        pending_logs = pending_logs.filter(direction=direction_filter)
    
    if dsc_search:
        pending_logs = pending_logs.filter(dsc_number__icontains=dsc_search)

    return render(request, 'delivery_collection.html', {
        'form': form, 
        'pending_logs': pending_logs,
        'direction_filter': direction_filter,
        'dsc_search': dsc_search
    })


@login_required
@role_required('admin')
def admin_initiation_requests(request):
    """View all DSC initiation requests with filtering."""
    initiation_requests = InOut.objects.filter(initiated=True).order_by('-initiation_date')
    
    # Apply simple filters
    dsc_number = request.GET.get('dsc_number')
    if dsc_number:
        initiation_requests = initiation_requests.filter(dsc_number__dsc_number__icontains=dsc_number)
    
    status = request.GET.get('status')
    if status == 'pending':
        initiation_requests = initiation_requests.filter(action_completed=False)
    elif status == 'completed':
        initiation_requests = initiation_requests.filter(action_completed=True)
    
    direction = request.GET.get('direction')
    if direction:
        initiation_requests = initiation_requests.filter(direction=direction)
    
    # Pagination
    paginator = Paginator(initiation_requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Add current direction to each request
    for request_obj in page_obj:
        latest_completed = (
            InOut.objects.filter(dsc_number=request_obj.dsc_number, action_completed=True)
            .order_by('-in_out_id')
            .first()
        )
        request_obj.current_direction = latest_completed.direction if latest_completed else "in"
    
    return render(request, 'admin_initiation_requests.html', {
        'initiation_requests': page_obj
    })


@login_required
@role_required('admin')  # ✅ Only admin can access
def admin_request_details(request, in_out_id):
    """View details of a specific DSC initiation request, including Delivery/Collection logs."""

    # ✅ Get the initiation request details
    request_obj = get_object_or_404(InOut, in_out_id=in_out_id)

    # ✅ Get latest completed direction (Actual status after delivery/collection)
    latest_completed = (
        InOut.objects.filter(dsc_number=request_obj.dsc_number, action_completed=True)
        .order_by('-in_out_id')
        .first()
    )
    current_direction = latest_completed.direction if latest_completed else "in"

    # ✅ Get all delivery/collection logs related to this DSC
    transaction_logs = InOut.objects.filter(dsc_number=request_obj.dsc_number, action_completed=True).order_by('-action_date', '-action_time')

    return render(request, 'admin_request_details.html', {
        'request_obj': request_obj,
        'current_direction': current_direction,
        'transaction_logs': transaction_logs
    })


@login_required
@role_required('admin')  # Ensure only admin can access
def dsc_list(request):
    today = date.today()
    
    # ✅ Filter by status
    filter_type = request.GET.get('filter', 'all')
    if filter_type == 'active':
        dscs = DSCMaster.objects.filter(expiry_date__gte=today, status="Active")
    elif filter_type == 'expired':
        dscs = DSCMaster.objects.filter(expiry_date__lt=today, status="Expired")
    else:
        dscs = DSCMaster.objects.all()

    # ✅ Search functionality
    query = request.GET.get('q', '')
    if query:
        dscs = dscs.filter(
            Q(dsc_number__icontains=query) |
            Q(full_name__icontains=query) |
            Q(email_id__icontains=query)
        )

    # ✅ Pagination: Show 10 records per page
    paginator = Paginator(dscs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "filter_type": filter_type,
        "query": query
    }
    return render(request, "dsc_list.html", context)


@login_required
@role_required('admin')
def export_dsc_list(request, format):
    """ Export DSC list as CSV or Excel """
    today = date.today()

    # ✅ Apply Filters
    filter_type = request.GET.get('filter', 'all')
    if filter_type == 'active':
        dscs = DSCMaster.objects.filter(expiry_date__gte=today, status="Active")
    elif filter_type == 'expired':
        dscs = DSCMaster.objects.filter(expiry_date__lt=today, status="Expired")
    else:
        dscs = DSCMaster.objects.all()

    # ✅ Apply Search
    query = request.GET.get('q', '')
    if query:
        dscs = dscs.filter(
            Q(dsc_number__icontains=query) |
            Q(full_name__icontains=query) |
            Q(email_id__icontains=query)
        )

    dsc_data = dscs.values('dsc_number', 'full_name', 'email_id', 'expiry_date', 'status')

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="dsc_list.csv"'
        writer = csv.writer(response)
        writer.writerow(['DSC Number', 'Full Name', 'Email ID', 'Expiry Date', 'Status'])

        for dsc in dsc_data:
            writer.writerow(dsc.values())

        return response

    elif format == 'excel':
        df = pd.DataFrame(dsc_data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="dsc_list.xlsx"'
        df.to_excel(response, index=False)

        return response

    return HttpResponse("Invalid format", status=400)



@login_required
@role_required('admin')
def manage_extra_fields(request):
    if request.method == "POST":
        field_name = request.POST.get("field_name")
        field_type = request.POST.get("field_type")
        dsc_type = "both" 

        # ✅ Prevent duplicate field names for the same DSC type
        if DSCExtraField.objects.filter(dsc_type=dsc_type, field_name=field_name).exists():
            messages.error(request, "❌ This field name already exists for this DSC type.")
        else:
            DSCExtraField.objects.create(dsc_type=dsc_type, field_name=field_name, field_type=field_type)
            messages.success(request, "✅ New field added successfully!")

        return redirect("manage_extra_fields")    

    extra_fields = DSCExtraField.objects.filter(Q(dsc_type="both") | Q(dsc_type=dsc_type))

    return render(request, "manage_extra_fields.html", {"extra_fields": extra_fields})




class ExtraFieldListView(ListView):
    model = DSCExtraField
    template_name = 'manage_extra_fields.html'
    context_object_name = 'extra_fields'
    ordering = ['field_name']

class ExtraFieldCreateView(CreateView):
    model = DSCExtraField
    template_name = 'extra_field_form.html'
    fields = ['dsc_type', 'field_name', 'field_type', 'required', 'label']  
    success_url = reverse_lazy('manage_extra_fields')

class ExtraFieldUpdateView(UpdateView):
    model = DSCExtraField
    template_name = 'extra_field_form.html'
    fields = ['dsc_type', 'field_name', 'field_type', 'required', 'label']
    success_url = reverse_lazy('manage_extra_fields')

class ExtraFieldDeleteView(DeleteView):
    model = DSCExtraField
    template_name = 'extra_field_confirm_delete.html'
    success_url = reverse_lazy('manage_extra_fields')




def send_dsc_expiry_notifications(expired_dscs, expiring_dscs):
    today = date.today()

    for dsc in expired_dscs:
        if dsc.email_id:  # ✅ Ensure email exists
            subject = f"⚠️ Your DSC ({dsc.dsc_number}) Has Expired!"
            message = f"""
Dear {dsc.full_name},

Your Digital Signature Certificate ({dsc.dsc_number}) expired on {dsc.expiry_date}.
Please renew it as soon as possible to continue using it.

If you have already renewed your DSC, please update the details.

Thank you,
DSC Management Team
"""
            send_mail(
                subject, message, 'admin@yourcompany.com', [dsc.email_id], fail_silently=False
            )

    for dsc in expiring_dscs:
        if dsc.email_id:
            subject = f"⏳ Your DSC ({dsc.dsc_number}) Will Expire Soon!"
            message = f"""
Dear {dsc.full_name},

Your Digital Signature Certificate ({dsc.dsc_number}) is set to expire on {dsc.expiry_date}.
Please ensure you renew it before the expiry date to avoid any disruption.

Thank you,
DSC Management Team
"""
            send_mail(
                subject, message, 'admin@yourcompany.com', [dsc.email_id], fail_silently=False
            )


@login_required
@role_required('admin')
def email_templates(request):
    templates = EmailTemplate.objects.all()
    return render(request, 'emails/email_templates.html', {'templates': templates})

@login_required
@role_required('admin')
def edit_template(request, template_type):
    template = EmailTemplate.objects.get(template_type=template_type)
    if request.method == 'POST':
        form = EmailTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, 'Template updated successfully!')
            return redirect('email_templates')
    else:
        form = EmailTemplateForm(instance=template)
    return render(request, 'emails/edit_template.html', {'form': form, 'template': template})

@login_required
@role_required('admin')
def send_manual_notifications(request, notification_type):
    today = date.today()
    
    if notification_type == 'expired':
        dscs = DSCMaster.objects.filter(expiry_date__lt=today, status="Expired")
        template = EmailTemplate.objects.get(template_type='expired')
    else:
        dscs = DSCMaster.objects.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=30),
            status="Active"
        )
        template = EmailTemplate.objects.get(template_type='expiring')
    
    # Send notifications
    send_dsc_expiry_notifications(
        expired_dscs=dscs if notification_type == 'expired' else None,
        expiring_dscs=dscs if notification_type == 'expiring' else None,
        template=template
    )
    
    messages.success(request, f'Notifications for {notification_type} DSCs sent successfully!')
    return redirect('email_templates')

from .notifications import send_dsc_expiry_notifications

@login_required
@role_required('admin')
def send_notifications(request, notification_type):
    today = date.today()
    
    try:
        if notification_type == 'expired':
            count = DSCMaster.objects.filter(
                expiry_date__lt=today, 
                status="Expired"
            ).count()
            # Call with named parameter
            send_dsc_expiry_notifications(notification_type='expired')
        else:
            count = DSCMaster.objects.filter(
                expiry_date__gte=today,
                expiry_date__lte=today + timedelta(days=30),
                status="Active"
            ).count()
            # Call with named parameter
            send_dsc_expiry_notifications(notification_type='expiring')
            
        messages.success(request, f'Notifications sent to {count} recipients!')
    except Exception as e:
        messages.error(request, f'Error sending notifications: {str(e)}')
    
    return redirect('email_templates')


@login_required
@user_passes_test(lambda u: u.role == 'admin')
@user_passes_test(lambda u: u.role == 'admin')
def remove_dscs(request):
    # Get all DSCs ordered by creation (newest first)
    all_dscs = DSCMaster.objects.all().order_by('-dsc_id')
    
    # Get the latest DSC ID
    latest_dsc_id = all_dscs.first().dsc_id if all_dscs.exists() else None
    
    # Add pagination
    paginator = Paginator(all_dscs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'remove_dsc.html', {
        'page_obj': page_obj,
        'latest_dsc_id': latest_dsc_id,
        'active_tab': 'remove_dsc'
    })

@login_required
@user_passes_test(lambda u: u.role == 'admin')
def confirm_delete_dsc(request, id):
    dsc = get_object_or_404(DSCMaster, dsc_id=id)
    latest_dsc = DSCMaster.objects.order_by('-dsc_id').first()
    
    return render(request, 'confirm_delete_dsc.html', {
        'dsc': dsc,
        'is_latest': dsc.dsc_id == latest_dsc.dsc_id if latest_dsc else False,
        'active_tab': 'remove_dsc'
    })

@login_required
@user_passes_test(lambda u: u.role == 'admin')
def delete_dsc(request, id):
    if request.method == 'POST':
        dsc = get_object_or_404(DSCMaster, dsc_id=id)
        latest_dsc_id = DSCMaster.objects.order_by('-dsc_id').values_list('dsc_id', flat=True).first()
        
        if dsc.dsc_id != latest_dsc_id:
            messages.error(request, 'You can only delete the most recently added DSC.')
            return redirect('remove_dscs')
            
        try:
            dsc_number = dsc.dsc_number
            dsc.delete()
            messages.success(request, f'DSC {dsc_number} has been deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting DSC: {str(e)}')
            
    return redirect('remove_dscs')
####################################################################general user################################################################
################################################################################################################################################


@login_required
@role_required('general-user')
def general_user_dashboard(request):
    today = date.today()

    # ✅ Fetch related entities in a single query
    dsc_entities = DSCEntity.objects.select_related("entity")
    
    # ✅ Fetch DSCs that are still "Expired" and "Pending Renewal"
    expired_pending_renewal = DSCMaster.objects.prefetch_related(
    Prefetch("dscentity_set", queryset=dsc_entities)
    ).filter(
        expiry_date__lt=today, 
        status="Expired"
    ).exclude(
        dsc_id__in=DSCRenewal.objects.filter(
            remarks__isnull=False  # ✅ Allow any remarks, not just "not renewing"
        ).values_list('dsc_number__dsc_id', flat=True)  
    )

    # ✅ Expired DSCs Marked as "Not Renewing"
    expired_not_renewing = DSCRenewal.objects.select_related("dsc_number").filter(
        remarks__isnull=False,  # ✅ Allow any remarks instead of just "not renewing"
        dsc_number__status="Expired"
    ).prefetch_related("dsc_number__dscentity_set__entity")


    # ✅ DSCs Expiring in Next 30 Days
    upcoming_expiring = DSCMaster.objects.filter(
        expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30), status="Active"
    )

    # ✅ DSCs Missing Additional Details (No Entry in DSCRenewal)
    missing_details = DSCMaster.objects.filter(
    
    Q(renewals__gst_reg_date__isnull=True) | 
    Q(renewals__mca_reg_date__isnull=True) | 
    Q(renewals__it_reg_date__isnull=True)
).distinct()

    context = {
        "expired_pending_renewal": expired_pending_renewal,  # ✅ FIXED
        "expired_not_renewing": expired_not_renewing,  
        "upcoming_expiring": upcoming_expiring,
        "missing_details": missing_details,
    }
    return render(request, "general_user/general_user_dashboard.html", context)



@login_required
@role_required('general-user')
def general_user_mark_not_renewing(request, dsc_number):
    """ Move a DSC from 'Expired Pending Renewal' to 'Expired Not Renewing' """
    dsc = get_object_or_404(DSCMaster, dsc_number=dsc_number)

    if request.method == "POST":
        remarks = request.POST.get("remarks", "").strip()

        if not remarks:
            messages.error(request, "❌ Remarks are required!")
            return redirect("general_user_dashboard")

        # ✅ Delete duplicate entries if they exist
        DSCRenewal.objects.filter(dsc_number=dsc).delete()

        # ✅ Create a new entry in DSCRenewal (marks it as Not Renewing)
        DSCRenewal.objects.create(dsc_number=dsc, remarks=remarks)

        messages.success(request, f"✅ DSC {dsc_number} moved to 'Not Renewing'.")
    
    return redirect("general_user_dashboard")



@login_required
@role_required('general-user')
def general_user_mark_pending_renewal(request, dsc_number):
    """ Move a DSC from 'Expired Not Renewing' back to 'Expired Pending Renewal' """
    renewal_entry = get_object_or_404(DSCRenewal, dsc_number__dsc_number=dsc_number)

    if request.method == "POST":
        # ✅ Delete the entry in DSCRenewal (removes it from 'Not Renewing')
        renewal_entry.delete()

        messages.success(request, f"✅ DSC {dsc_number} moved back to 'Pending Renewal'.")
    
    return redirect("general_user_dashboard")

@login_required
@role_required('general-user')
def general_user_renewal(request):
    search_query = request.GET.get('search', '')
    dsc_number = request.GET.get('dsc_number')
    
    # Start with all DSCs ordered by DSC number
    dsc_numbers = DSCMaster.objects.all().order_by('dsc_number')
    
    # Apply search filter if provided
    if search_query:
        dsc_numbers = dsc_numbers.filter(
            Q(dsc_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(pan_no__icontains=search_query) |
            Q(email_id__icontains=search_query) |
            Q(phone_no__icontains=search_query)
        )
    
    dsc_master = None
    dsc_form = None
    doc_form = None
    documents = None
    extra_fields_with_values = []

    if request.method == 'POST':
        dsc_number = request.POST.get('dsc_number')
        if not dsc_number:
            messages.error(request, "❌ Please select a DSC number.")
            return redirect('general_user_renewal')

        dsc_master = get_object_or_404(DSCMaster, dsc_number=dsc_number)
        old_password = dsc_master.password  

        # Capture data before renewal for history
        previous_data = {
            'dsc_number': dsc_master.dsc_number,
            'full_name': dsc_master.full_name,
            'issued_date': dsc_master.issued_date.isoformat(),
            'expiry_date': dsc_master.expiry_date.isoformat() if dsc_master.expiry_date else None,
            'license_period': dsc_master.license_period.no_of_years if dsc_master.license_period else None,
            'password': dsc_master.password,
            'pan_no': dsc_master.pan_no,
            'dsc_class': dsc_master.dsc_class.class_name if dsc_master.dsc_class else None,
            'email_id': dsc_master.email_id,
            'phone_no': dsc_master.phone_no,
            'issuing_auth': dsc_master.issuing_auth.auth_name if dsc_master.issuing_auth else None,
            'ref_name': dsc_master.ref_name,
            'ref_contact': dsc_master.ref_contact,
            'type': dsc_master.type,
            'remarks': dsc_master.remarks,
            'extra_fields': dsc_master.extra_fields,
        }
        
        # Capture document info before renewal
        current_docs = Docs.objects.filter(dsc_number=dsc_master).first()
        previous_docs = {
            'aadhar': current_docs.aadhar_path.url if current_docs and current_docs.aadhar_path else None,
            'pan': current_docs.pan_path.url if current_docs and current_docs.pan_path else None,
            'pp': current_docs.pp_path.url if current_docs and current_docs.pp_path else None,
        }

        # Get extra fields for this DSC type
        extra_fields = DSCExtraField.objects.filter(
            Q(dsc_type=dsc_master.dsc_type.type_name.lower()) | Q(dsc_type="both"))

        # Choose the correct form based on DSC type
        if dsc_master.dsc_type.type_name.lower() == "internal":
            DscFormClass = InternalDSCForm
        else:
            DscFormClass = ExternalDSCForm  

        dsc_form = DscFormClass(request.POST, instance=dsc_master)
        documents = Docs.objects.filter(dsc_number=dsc_master)
        doc_form = DocsForm(request.POST, request.FILES)

        if dsc_form.is_valid() and doc_form.is_valid():
            # Process extra fields - update both storage locations
            extra_data = {}
            for field in extra_fields:
                field_value = request.POST.get(field.field_name)
                if field_value:
                    if field.field_type == "boolean":
                        field_value = True if field_value == "on" else False
                    
                    # Update JSON field
                    extra_data[field.field_name] = field_value
                    
                    # Update or create DSCExtraData record
                    DSCExtraData.objects.update_or_create(
                        dsc_number=dsc_master,
                        field_name=field.field_name,
                        defaults={'value': str(field_value)}
                    )
            
            dsc_master.extra_fields = extra_data

            # Handle password
            new_password = dsc_form.cleaned_data.get('password')
            if not new_password:
                dsc_master.password = old_password

            # Calculate new expiry date
            new_issued_date = dsc_form.cleaned_data.get('issued_date')
            new_license_period = dsc_form.cleaned_data.get('license_period')
            new_expiry_date = new_issued_date + timedelta(days=new_license_period.no_of_years * 365)
            dsc_master.expiry_date = new_expiry_date
            dsc_master.status = "Active" if new_expiry_date > now().date() else "Expired"
            dsc_master.save()

            # Handle document updates
            new_docs = {}
            if current_docs:
                if 'aadhar_path' in request.FILES:
                    current_docs.aadhar_path = request.FILES['aadhar_path']
                if 'pan_path' in request.FILES:
                    current_docs.pan_path = request.FILES['pan_path']
                if 'pp_path' in request.FILES:
                    current_docs.pp_path = request.FILES['pp_path']
                current_docs.save()
                
                new_docs = {
                    'aadhar': current_docs.aadhar_path.url if current_docs.aadhar_path else None,
                    'pan': current_docs.pan_path.url if current_docs.pan_path else None,
                    'pp': current_docs.pp_path.url if current_docs.pp_path else None,
                }
            else:
                doc_instance = doc_form.save(commit=False)
                doc_instance.dsc_number = dsc_master
                doc_instance.save()
                new_docs = {
                    'aadhar': doc_instance.aadhar_path.url if doc_instance.aadhar_path else None,
                    'pan': doc_instance.pan_path.url if doc_instance.pan_path else None,
                    'pp': doc_instance.pp_path.url if doc_instance.pp_path else None,
                }

            # Prepare new data for history
            new_data = {
                'dsc_number': dsc_form.cleaned_data.get('dsc_number'),
                'full_name': dsc_form.cleaned_data.get('full_name'),
                'issued_date': new_issued_date.isoformat(),
                'expiry_date': new_expiry_date.isoformat(),
                'license_period': new_license_period.no_of_years,
                'password': new_password if new_password else old_password,
                'pan_no': dsc_form.cleaned_data.get('pan_no'),
                'dsc_class': dsc_form.cleaned_data.get('dsc_class').class_name if dsc_form.cleaned_data.get('dsc_class') else None,
                'email_id': dsc_form.cleaned_data.get('email_id'),
                'phone_no': dsc_form.cleaned_data.get('phone_no'),
                'issuing_auth': dsc_form.cleaned_data.get('issuing_auth').auth_name if dsc_form.cleaned_data.get('issuing_auth') else None,
                'ref_name': dsc_form.cleaned_data.get('ref_name'),
                'ref_contact': dsc_form.cleaned_data.get('ref_contact'),
                'type': dsc_form.cleaned_data.get('type'),
                'remarks': dsc_form.cleaned_data.get('remarks'),
                'extra_fields': extra_data,
            }

            # Create renewal history record
            DSCRenewalHistory.objects.create(
                dsc=dsc_master,
                renewed_by=request.user,
                previous_data=previous_data,
                new_data=new_data,
                previous_documents=previous_docs,
                new_documents=new_docs
            )

            # Remove DSC from "Not Renewing" category if mistakenly added
            DSCRenewal.objects.filter(dsc_number=dsc_master, remarks__isnull=False).delete()

            messages.success(request, "✅ DSC renewed successfully!")
            return redirect('general_user_renewal'  )
        else:
            messages.error(request, "❌ Please fill in all required fields.")
            print("Form Errors (dsc_form):", dsc_form.errors)
            print("Form Errors (doc_form):", doc_form.errors)

    # Handle GET request
    dsc_number = request.GET.get('dsc_number')
    if dsc_number:
        dsc_master = get_object_or_404(DSCMaster, dsc_number=dsc_number)
        
        # Get extra fields for this DSC type
        extra_fields = DSCExtraField.objects.filter(
            Q(dsc_type=dsc_master.dsc_type.type_name.lower()) | Q(dsc_type="both"))
        
        # Prepare extra fields with values - check both sources
        extra_fields_with_values = []
        for field in extra_fields:
            # First check JSON field
            field_value = dsc_master.extra_fields.get(field.field_name) if dsc_master.extra_fields else None
            
            # If not found in JSON, check DSCExtraData
            if field_value is None:
                try:
                    extra_data = DSCExtraData.objects.get(
                        dsc_number=dsc_master,
                        field_name=field.field_name
                    )
                    field_value = extra_data.value
                    # Update JSON field for consistency
                    if not dsc_master.extra_fields:
                        dsc_master.extra_fields = {}
                    dsc_master.extra_fields[field.field_name] = field_value
                    dsc_master.save()
                except DSCExtraData.DoesNotExist:
                    field_value = ''
            
            extra_fields_with_values.append({
                'field_name': field.field_name,
                'display_name': field.label or field.field_name,
                'value': field_value,
                'field_type': field.field_type,
                'is_required': field.required,
                'choices': field.get_FIELD_TYPE_CHOICES_display() if field.field_type == 'select' else []
            })
        
        # Choose the correct form dynamically
        if dsc_master.dsc_type.type_name.lower() == "internal":
            DscFormClass = InternalDSCForm
        else:
            DscFormClass = ExternalDSCForm
            
        dsc_form = DscFormClass(instance=dsc_master) 
        dsc_form.fields['password'].widget.attrs['value'] = dsc_master.password  # Pre-fill password
        documents = Docs.objects.filter(dsc_number=dsc_master)
        doc_form = DocsForm()  # Form for uploading new documents
        
    return render(request, 'general_user/renewal.html', {
        'dsc_master': dsc_master,
        'dsc_numbers': dsc_numbers,
        'dsc_form': dsc_form,
        'doc_form': doc_form,
        'documents': documents,
        'extra_fields_with_values': extra_fields_with_values,
        'history_count': DSCRenewalHistory.objects.filter(dsc=dsc_master).count() if dsc_master else 0,
    })



@login_required
@role_required('general-user')
def general_user_search_dsc(request):
    query = request.GET.get('q', '').strip()  # Get the search query
    search_type = request.GET.get('type', 'name')  # Determine search type

    results = []    
    if query:
        if search_type == 'name':
            results = DSCMaster.objects.filter(full_name__icontains=query)
        elif search_type == 'entity':
            results = DSCMaster.objects.filter(dscentity_set__entity__entity_name__icontains=query).distinct()
        elif search_type == 'pan':
            results = DSCMaster.objects.filter(pan_no__icontains=query)
        elif search_type == 'expiry_date':
            results = DSCMaster.objects.filter(expiry_date__icontains=query)
        elif search_type == 'status':
            results = DSCMaster.objects.filter(status__icontains=query)  # Active / Expired

    # ✅ Attach latest **completed** direction to each DSC result
    for dsc in results:
        latest_completed_inout = (
            InOut.objects.filter(dsc_number=dsc, action_completed=True)
            .order_by('-in_out_id')
            .first()
        )
        dsc.latest_direction = latest_completed_inout.direction if latest_completed_inout else "in"

    return render(request, 'general_user/search_results.html', {
        'query': query,
        'search_type': search_type,
        'results': results
    })


@login_required
@role_required('general-user')
def general_user_dsc_display(request, dsc_number):
    dsc_master = get_object_or_404(DSCMaster, dsc_number=dsc_number)

    # ✅ Get the latest completed transaction (Actual last known direction)
    latest_completed = (
        InOut.objects.filter(dsc_number=dsc_master, action_completed=True)
        .order_by('-in_out_id')
        .first()
    )

    # ✅ Get the latest initiated transaction (Ongoing initiation)
    latest_initiated = (
        InOut.objects.filter(dsc_number=dsc_master, initiated=True, action_completed=False)
        .order_by('-in_out_id')
        .first()
    )

    # ✅ Correct Current Direction Logic
    if latest_completed:
        current_direction = latest_completed.direction  # ✅ Use last **completed** direction
    else:
        current_direction = "in"  # ✅ Default for new DSCs

    # ✅ Fetch related data
    documents = Docs.objects.filter(dsc_number=dsc_master)
    entities = (
        DSCEntity.objects.filter(dsc_number=dsc_master)
        .values_list('entity__entity_name', flat=True)
    )

    # Process additional fields - check both JSON field and DSCExtraData
    additional_fields = []
    extra_fields = DSCExtraField.objects.filter(
        Q(dsc_type=dsc_master.dsc_type.type_name.lower()) | Q(dsc_type="both")
    )
    
    for field in extra_fields:
        # First check JSON field
        value = dsc_master.extra_fields.get(field.field_name) if dsc_master.extra_fields else None
        
        # If not found in JSON, check DSCExtraData
        if value is None:
            try:
                extra_data = DSCExtraData.objects.get(
                    dsc_number=dsc_master,
                    field_name=field.field_name
                )
                value = extra_data.value
                # Update JSON field for consistency
                if not dsc_master.extra_fields:
                    dsc_master.extra_fields = {}
                dsc_master.extra_fields[field.field_name] = value
                dsc_master.save()
            except DSCExtraData.DoesNotExist:
                value = None
        
        additional_fields.append({
            'name': field.label or field.field_name.replace('_', ' ').title(),
            'value': value if value not in [None, ""] else "Not provided"
        })
    
    print("\nDEBUG - Fields to Display:", additional_fields)  # Verify before rendering
    
    return render(request, 'general_user/dscDisplay.html', {
        'dsc_master': dsc_master,
        'entities': entities,
        'documents': documents,
        'current_direction': current_direction,  # ✅ Pass to template
        'additional_fields': additional_fields,
    })





@login_required
@role_required('general-user')
def general_user_map_dsc_entity(request):
    """Handles mapping of DSCs to Entities and displays mapped records."""
    dsc_numbers = DSCMaster.objects.all().order_by('dsc_number')

    if request.method == 'POST':
        form = DSCEntityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Entity mapped to DSC successfully!")
            return redirect('general_user_map_dsc_entity')
        else:
            messages.error(request, "❌ Error in mapping DSC to Entity.")
    else:
        form = DSCEntityForm()

    selected_entity_id = request.GET.get("entity_id")  # Get selected entity from dropdown
    entities = Entity.objects.all()  # Fetch all available entities

    # ✅ Fetch only MAPPED DSC-Entities
    mappings = DSCEntity.objects.select_related("dsc_number", "entity")

    # ✅ Apply entity filter if selected
    if selected_entity_id:
        mappings = mappings.filter(entity_id=selected_entity_id)

    return render(request, "general_user/map_dsc_entity.html", {
        "form": form,
        "entities": entities,
        "mappings": mappings,
        "selected_entity_id": selected_entity_id,
        "dsc_numbers": dsc_numbers,
    })


@login_required
@role_required('general-user')
def general_user_delete_mapping(request, mapping_id):
    """Deletes a DSC-Entity mapping."""
    mapping = get_object_or_404(DSCEntity, id=mapping_id)
    mapping.delete()
    messages.success(request, "✅ Mapping deleted successfully!")
    return redirect('general_user_map_dsc_entity')



@login_required
@role_required('general-user')
def general_user_internal_newentry(request):
    dsc_type = "internal"
    form_new = InternalDSCForm()
    form_docs = DocsForm(dsc_type='Internal')
    existing_files = {}
    extra_fields = DSCExtraField.objects.filter(Q(dsc_type="both") | Q(dsc_type=dsc_type)).order_by('id')
  # ✅ Fetch all extra fields for both types
    extra_data = {}  # Store extra field values

    # Auto-generate next DSC number and make it readonly in the view
    # Get last internal DSC with internal type
    try:
        internal_type = get_object_or_404(Type, type_name='Internal')
        last_internal_dsc = DSCMaster.objects.filter(dsc_type=internal_type).order_by('-dsc_id').first()
        
        if last_internal_dsc and last_internal_dsc.dsc_number:
            # Extract numeric part if the format is like "INT-001"
            if '-' in last_internal_dsc.dsc_number:
                prefix, number = last_internal_dsc.dsc_number.split('-')
                new_number = int(number) + 1
                next_dsc_number = f"{prefix}-{new_number:03d}"
            else:
                # If it's just a number, increment it
                next_dsc_number = str(int(last_internal_dsc.dsc_number) + 1)
        else:
            # First entry
            next_dsc_number = "1500"
            
        # Initialize the field with auto-generated value
        form_new.fields['dsc_number'].initial = next_dsc_number
        # Make the field readonly
        form_new.fields['dsc_number'].widget.attrs['readonly'] = True
    except Exception as e:
        # Handle any exceptions
        pass


    if request.method == 'POST':

        # Get the generated DSC number from the form
        generated_dsc_number = request.POST.get('dsc_number')
        form_new = InternalDSCForm(request.POST)
        form_docs = DocsForm(request.POST, request.FILES, dsc_type='Internal')

        
        # ✅ Handle extra fields
    # for field in extra_fields:
    #     extra_data[field.field_name] = request.POST.get(field.field_name, '')


        # Make sure the field is readonly in POST too
        form_new.fields['dsc_number'].widget.attrs['readonly'] = True

        if form_new.is_valid() and form_docs.is_valid():
            form_new_instance = form_new.save(commit=False)

            # Ensure we use the generated DSC number
            form_new_instance.dsc_number = generated_dsc_number

            # ✅ Set the logged-in user as user_id
            form_new_instance.user_id = request.user  # Assign the logged-in user directly

            # ✅ Automatically set DSC Type to Internal
            internal_type = get_object_or_404(Type, type_name='Internal')  
            form_new_instance.dsc_type = internal_type

            # ✅ Handle expiry status
            issued_date = form_new.cleaned_data['issued_date']
            license_period = form_new.cleaned_data['license_period'].no_of_years
            expiry_date = issued_date.replace(year=issued_date.year + license_period)

            form_new_instance.status = "Expired" if expiry_date < now().date() else "Active"
            form_new_instance.save()

            # Initialize extra_fields as empty dict if None
            if not hasattr(form_new_instance, 'extra_fields'):
                form_new_instance.extra_fields = {}
            
            # Save extra fields to both DSCExtraData AND DSCMaster.extra_fields
            for field in extra_fields:
                field_value = request.POST.get(field.field_name)
                if field_value:
                    # Save to DSCExtraData
                    DSCExtraData.objects.update_or_create(
                        dsc_number=form_new_instance,
                        field_name=field.field_name,
                        defaults={'value': field_value}
                    )
                    
                    # Also save to DSCMaster.extra_fields JSON field
                    if field.field_type == "boolean":
                        field_value = True if field_value == "on" else False
                    form_new_instance.extra_fields[field.field_name] = field_value

            # Save the instance after updating extra_fields
            form_new_instance.save()                    
            
            
            # ✅ Save uploaded documents
            form_docs_instance = form_docs.save(commit=False)
            form_docs_instance.dsc_number = form_new_instance
            form_docs_instance.save()

            # ✅ Automatically Create InOut Entry
            InOut.objects.create(
                dsc_number=form_new_instance,
                direction="in",
                initiated=False,
                action_completed=False
            )

            messages.success(request, "✅ Internal DSC entry added successfully!")
            return redirect('general_user_map_dsc_entity')

        else:
            messages.error(request, "❌ Please fill all required fields for Internal DSC.")
            print("Form Errors (form_new):", form_new.errors)
            print("Form Errors (form_docs):", form_docs.errors)


            # ✅ Store uploaded file names to keep them after a validation failure
            for field in ['aadhar_path', 'pan_path', 'pp_path']:
                if field in request.FILES:
                    existing_files[field] = request.FILES[field].name  # ✅ Store filename

    return render(request, 'general_user/internal_newentry.html', {
        'form_new': form_new, 
        'form_docs': form_docs,
        'existing_files': existing_files,  # ✅ Pass retained file names to template
        'extra_fields': extra_fields,  # ✅ Pass extra fields to template
    })



@login_required
@role_required('general-user')
def general_user_external_newentry(request):
    dsc_type = "external"
    form_new = ExternalDSCForm()
    form_docs = DocsForm(dsc_type='External')
    extra_fields = DSCExtraField.objects.filter(Q(dsc_type="both") | Q(dsc_type=dsc_type)).order_by('id')
  # ✅ Fetch all extra fields for both types
    extra_data = {}  # Store extra field values

    
    if request.method == 'POST':
        form_new = ExternalDSCForm(request.POST)
        form_docs = DocsForm(request.POST, request.FILES, dsc_type='External')  # ✅ Pass dsc_type


        if form_new.is_valid() and form_docs.is_valid():
            form_new_instance = form_new.save(commit=False)
            
            # ✅ Set the logged-in user as user_id
            form_new_instance.user_id = request.user  

            # ✅ Automatically set DSC Type to External
            external_type = get_object_or_404(Type, type_name='External')
            form_new_instance.dsc_type = external_type

            # ✅ Handle optional fields correctly
            form_new_instance.dsc_class = form_new.cleaned_data.get('dsc_class') or None
            form_new_instance.issuing_auth = form_new.cleaned_data.get('issuing_auth') or None
            form_new_instance.type = form_new.cleaned_data.get('type') or None

            # ✅ Handle expiry status
            issued_date = form_new.cleaned_data['issued_date']
            license_period = form_new.cleaned_data['license_period'].no_of_years
            expiry_date = issued_date.replace(year=issued_date.year + license_period)

            form_new_instance.status = "Expired" if expiry_date < now().date() else "Active"
            form_new_instance.save()

            # Initialize extra_fields if None
            if not hasattr(form_new_instance, 'extra_fields'):
                form_new_instance.extra_fields = {}
            
            # Save extra fields to both storage locations
            for field in extra_fields:
                field_value = request.POST.get(field.field_name)
                if field_value:
                    # Save to DSCExtraData
                    DSCExtraData.objects.update_or_create(
                        dsc_number=form_new_instance,
                        field_name=field.field_name,
                        defaults={'value': field_value}
                    )
                    
                    # Also save to DSCMaster.extra_fields JSON field
                    if field.field_type == "boolean":
                        field_value = True if field_value == "on" else False
                    form_new_instance.extra_fields[field.field_name] = field_value
            
            form_new_instance.save()
                     
            # Save documents                     
            form_docs_instance = form_docs.save(commit=False)
            form_docs_instance.dsc_number = form_new_instance
            form_docs_instance.save()

            messages.success(request, "✅ External DSC entry added successfully!")
            return redirect('general_user_map_dsc_entity')

        else:
            print("❌ External DSC Form Errors:", form_new.errors)
            print("❌ Documents Form Errors:", form_docs.errors)
            messages.error(request, "❌ Please fill in all required fields for External DSC. Please check the form. There are some errors.")

    else:
        form_new = ExternalDSCForm()
        form_docs = DocsForm(dsc_type='external')  # ✅ Pass dsc_type

    return render(request, 'general_user/external_newentry.html', {'form_new': form_new, 'form_docs': form_docs,'extra_fields': extra_fields})# ✅ Pass extra fields to template





@login_required
@role_required('general-user')
def general_user_initiate_dsc(request):
    """Handles initiation of DSC movement."""
    if request.method == 'POST':
        form = InitiationForm(request.POST)

        if form.is_valid():
            initiation = form.save(commit=False)
            initiation.initiated_by = request.user
            initiation.initiated = True  # ✅ Mark as initiated

            # ✅ Auto-fill requester_name if "Self" was selected
            if request.POST.get('requester_type_value') == "self":
                initiation.requester_name = request.user.username  # Assign logged-in user's name

            if initiation.initiation_date is None:
                initiation.initiation_date = now()

            initiation.save()
            messages.success(request, f"✅ DSC {initiation.dsc_number} has been initiated.")
            return redirect('general_user_initiate_dsc')
        else:
            messages.error(request, "❌ Form submission failed. Please check the fields.")

    else:
        form = InitiationForm()

    # ✅ Fetch **only pending** initiated logs (not completed ones)
    initiated_logs = InOut.objects.filter(initiated=True, action_completed=False, initiated_by__isnull=False).select_related('dsc_number')


    # ✅ Fetch last completed direction for each DSC
    for log in initiated_logs:
        last_completed = (
            InOut.objects.filter(dsc_number=log.dsc_number, action_completed=True)
            .order_by('-action_date', '-action_time')
            .first()
        )
        log.last_completed_direction = last_completed.direction if last_completed else None

    dsc_list = DSCMaster.objects.all()

    return render(request, 'general_user/initiation.html', {
        'form': form,
        'initiated_logs': initiated_logs,
        'dsc_list': dsc_list,
    })



@login_required
@role_required('general-user')
def general_user_delivery_collection(request):
    """Handles marking of DSC as Delivered or Collected."""
    if request.method == 'POST':
        log_id = request.POST.get('log_id')

        if not log_id:
            messages.error(request, "❌ No DSC was selected.")
            return redirect('general_user_delivery_collection')

        log = get_object_or_404(InOut, in_out_id=log_id)
        form = DeliveryCollectionForm(request.POST, instance=log)

        if form.is_valid():
            delivery = form.save(commit=False)
            delivery.action_completed = True  # ✅ Mark as completed

            if not delivery.action_date:
                delivery.action_date = now().date()

            if not delivery.action_time:
                delivery.action_time = now().time()

            # ✅ Only change direction after successful delivery/collection
            if log.direction == "in":  
                delivery.direction = "out"
            else:
                delivery.direction = "in"

            delivery.save()
            messages.success(request, f"✅ DSC {delivery.dsc_number} has been successfully delivered/collected.")
            return redirect('general_user_delivery_collection')
        else:
            messages.error(request, "❌ There was an error in form submission.")

    else:
        form = DeliveryCollectionForm()

    
        # ✅ Fetch pending logs and ensure each log has its initiated direction
    pending_logs = InOut.objects.filter(initiated=True, action_completed=False, initiated_by__isnull=False).annotate(
        initiated_direction=F('direction')  # ✅ Get the direction at the time of initiation
    )
    # ✅ Add filtering functionality
    direction_filter = request.GET.get('direction')
    dsc_search = request.GET.get('dsc_search')

    if direction_filter:
        pending_logs = pending_logs.filter(direction=direction_filter)
    
    if dsc_search:
        pending_logs = pending_logs.filter(dsc_number__icontains=dsc_search)


    return render(request, 'general_user/delivery_collection.html', { 
        'form': form, 
        'pending_logs': pending_logs,
        'direction_filter': direction_filter,
        'dsc_search': dsc_search})


@login_required
@role_required('general-user')  # ✅ Only admins can access
def general_user_initiation_requests(request):
    """View all DSC initiation requests with initiated and current direction."""

    # ✅ Only show initiated DSCs (i.e., not newly added DSCs)
    initiation_requests = InOut.objects.filter(initiated=True).order_by('-initiation_date')

    # Apply simple filters
    dsc_number = request.GET.get('dsc_number')
    if dsc_number:
        initiation_requests = initiation_requests.filter(dsc_number__dsc_number__icontains=dsc_number)
    
    status = request.GET.get('status')
    if status == 'pending':
        initiation_requests = initiation_requests.filter(action_completed=False)
    elif status == 'completed':
        initiation_requests = initiation_requests.filter(action_completed=True)
    
    direction = request.GET.get('direction')
    if direction:
        initiation_requests = initiation_requests.filter(direction=direction)
    
    # Pagination
    paginator = Paginator(initiation_requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


# Add current direction to each request
    for request_obj in page_obj:
        latest_completed = (
            InOut.objects.filter(dsc_number=request_obj.dsc_number, action_completed=True)
            .order_by('-in_out_id')
            .first()
        )
        request_obj.current_direction = latest_completed.direction if latest_completed else "in"


    return render(request, 'general_user/user_initiation_requests.html', {
        'initiation_requests': page_obj
    })


@login_required
@role_required('general-user')  # ✅ Only admin can access
def general_user_request_details(request, in_out_id):
    """View details of a specific DSC initiation request, including Delivery/Collection logs."""

    # ✅ Get the initiation request details
    request_obj = get_object_or_404(InOut, in_out_id=in_out_id)

    # ✅ Get latest completed direction (Actual status after delivery/collection)
    latest_completed = (
        InOut.objects.filter(dsc_number=request_obj.dsc_number, action_completed=True)
        .order_by('-in_out_id')
        .first()
    )
    current_direction = latest_completed.direction if latest_completed else "in"

    # ✅ Get all delivery/collection logs related to this DSC
    transaction_logs = InOut.objects.filter(dsc_number=request_obj.dsc_number, action_completed=True).order_by('-action_date', '-action_time')

    return render(request, 'general_user/user_request_details.html', {
        'request_obj': request_obj,
        'current_direction': current_direction,
        'transaction_logs': transaction_logs
    })



@login_required
@role_required('general-user')
def general_user_usage_logs(request):
    if request.method == "POST":
        print("Received POST data:", request.POST)  # ✅ Debugging

        # ✅ Get POST data
        form_other = request.POST.get("form_other", "").strip()
        platform_other = request.POST.get("platform_other", "").strip()
        form_id = request.POST.get("form_id")
        platform_id = request.POST.get("platform_id")

        # ✅ If "Other" is selected, create a new Form object
        if form_id == "other":
            if not form_other:
                messages.error(request, "❌ Please enter a form name.")
                return redirect("general_user_usage_logs")
            form_obj, created = Forms.objects.get_or_create(form_name=form_other)
            form_id = form_obj.pk  # ✅ Use the new Form ID

        # ✅ If "Other" is selected, create a new Platform object
        if platform_id == "other":
            if not platform_other:
                messages.error(request, "❌ Please enter a platform name.")
                return redirect("general_user_usage_logs")
            platform_obj, created = Platform.objects.get_or_create(platform_name=platform_other)
            platform_id = platform_obj.pk  # ✅ Use the new Platform ID

        # ✅ Modify request.POST to include new IDs
        mutable_post = request.POST.copy()
        mutable_post["form_id"] = form_id
        mutable_post["platform_id"] = platform_id

        form = UsageLogsForm(mutable_post, request=request)

        if form.is_valid():
            print("✅ Form is valid")  # ✅ Debugging
            usage_log = form.save(commit=False)

            # ✅ Assign current user
            usage_log.requested_by = request.user
            usage_log.save()

            messages.success(request, "✅ Usage log submitted successfully! Waiting for approval.")
            return redirect("general_user_usage_logs")

        else:
            print("❌ Form errors:", form.errors)  # ✅ Debugging
            messages.error(request, "❌ Error in submission. Please check the form.")

    else:
        form = UsageLogsForm(request=request)

    usage_logs_data = UsageLogs.objects.all()
        # Filtering logic
    usage_logs_data = UsageLogs.objects.all().select_related(
        'dsc_id', 'form_id', 'platform_id', 'requested_by', 'approved_by_user'
    ).order_by('-date_of_usage')

    search_query = request.GET.get('search')
    if search_query:
        usage_logs_data = usage_logs_data.filter(
            Q(dsc_id__dsc_number__icontains=search_query) |
            Q(form_id__form_name__icontains=search_query) |
            Q(platform_id__platform_name__icontains=search_query) |
            Q(requested_by__username__icontains=search_query)
        )

    status_filter = request.GET.get('status')
    if status_filter:
        usage_logs_data = usage_logs_data.filter(approval_status=status_filter)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        usage_logs_data = usage_logs_data.filter(
            date_of_usage__gte=start_date,
            date_of_usage__lte=end_date
        )


    paginator = Paginator(usage_logs_data, 15)  
    page_number = request.GET.get('page')
    usage_logs_data = paginator.get_page(page_number)

    return render(request, "general_user/usagelogs.html", {
        "form": form,
        "usage_logs_data": usage_logs_data,
    })

# Delete Usage Log
@login_required
@role_required('general-user')
def general_user_delete_usage_log(request, log_id):
    usage_log = get_object_or_404(UsageLogs, log_id=log_id)

    usage_log.delete()
    messages.success(request, "Usage log deleted successfully!")
    return redirect('general_user_usage_logs')


@login_required
@role_required('general-user')
def general_user_additional_details(request):
    """Manage Additional Details for DSCs"""
    dsc_numbers = DSCMaster.objects.all()
    additional_form = None
    dsc_master = None
    previous_additional_details = None
    extra_fields_data = []

    if request.method == 'POST':
        dsc_number = request.POST.get('dsc_number') or request.POST.get('dsc_number_hidden')

        if not dsc_number:
            messages.error(request, "❌ No DSC selected. Please choose a DSC.")
            return redirect('general_user_additional_details')

        try:
            dsc_master = DSCMaster.objects.get(dsc_number=dsc_number)
        except DSCMaster.DoesNotExist:
            messages.error(request, f"❌ No DSC found with number {dsc_number}. Please try again.")
            return redirect('general_user_additional_details')

        previous_additional_details = DSCRenewal.objects.filter(dsc_number=dsc_master).order_by('-additional_id')
        latest_additional_instance = previous_additional_details.first()

        additional_form = DSCRenewalForm(request.POST, instance=latest_additional_instance)

        if additional_form.is_valid():
            additional_instance = additional_form.save(commit=False)
            additional_instance.dsc_number = dsc_master

            # Save extra fields to JSON field with proper serialization
            additional_instance.extra_fields = {}
            for field in RenewalExtraField.objects.all():
                field_name = field.field_name
                if field_name in additional_form.cleaned_data:
                    value = additional_form.cleaned_data[field_name]
                    
                    # Handle different field types
                    if field.field_type == "boolean":
                        value = bool(value)
                    elif field.field_type == "date" and value:
                        if isinstance(value, datetime_date):
                            value = value.isoformat()  # Convert date to string
                    elif value is None:
                        continue
                        
                    additional_instance.extra_fields[field_name] = value

            additional_instance.save()
            messages.success(request, "✅ Additional details updated successfully!")
            return redirect('general_user_additional_details')

    dsc_number = request.GET.get('dsc_number')
    if dsc_number:
        try:
            dsc_master = DSCMaster.objects.get(dsc_number=dsc_number)
            previous_additional_details = DSCRenewal.objects.filter(dsc_number=dsc_master).order_by('-additional_id')
            latest_additional_instance = previous_additional_details.first()

            if latest_additional_instance:
                additional_form = DSCRenewalForm(instance=latest_additional_instance)
                # Deserialize dates from JSON
                if latest_additional_instance.extra_fields:
                    for field in RenewalExtraField.objects.filter(field_type='date'):
                        if field.field_name in latest_additional_instance.extra_fields:
                            date_str = latest_additional_instance.extra_fields[field.field_name]
                            if date_str:
                                try:
                                    # Fixed: Use datetime_date.fromisoformat()
                                    additional_form.initial[field.field_name] = datetime_date.fromisoformat(date_str)
                                except (ValueError, TypeError):
                                    pass
            else:
                additional_form = DSCRenewalForm()

            # Prepare extra fields data for template
            if additional_form:
                for field in RenewalExtraField.objects.all():
                    try:
                        extra_fields_data.append({
                            'name': field.field_name,
                            'field': additional_form[field.field_name],
                            'label': field.label or field.field_name.replace('_', ' ').title(),
                            'required': field.required,
                            'help_text': field.help_text,
                            'field_type': field.field_type  # Add field type to context
                        })
                    except KeyError:
                        continue

        except DSCMaster.DoesNotExist:
            messages.error(request, f"❌ No DSC found with number {dsc_number}. Please try again.")
            return redirect('general_user_additional_details')

    return render(request, 'general_user/additional_details.html', {
        'dsc_numbers': dsc_numbers,
        'additional_form': additional_form,
        'selected_dsc': dsc_master,
        'previous_additional_details': previous_additional_details,
        'extra_fields_data': extra_fields_data,
    })



##########################################################################################################################################
########################################################### APPROEVER ########################################################################
#################################################################################################################################


from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import UsageLogs
from .decorators import role_required

@login_required
@role_required('approver')
def approver_dashboard(request):
    """ Dashboard for pending approvals only """
    # Get base queryset
    pending_requests = UsageLogs.objects.filter(approval_status="pending")\
        .select_related('dsc_id', 'requested_by')\
        .order_by('-date_of_usage')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    requester_filter = request.GET.get('requester', '')
    
    # Apply filters
    if search_query:
        pending_requests = pending_requests.filter(
            Q(dsc_id__dsc_number__icontains=search_query) |
            Q(dsc_id__full_name__icontains=search_query) |
            Q(form_name__icontains=search_query) |
            Q(platform_name__icontains=search_query)
        )
    
    if requester_filter:
        pending_requests = pending_requests.filter(requested_by__username=requester_filter)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(pending_requests, 10)
    
    try:
        pending_requests = paginator.page(page)
    except PageNotAnInteger:
        pending_requests = paginator.page(1)
    except EmptyPage:
        pending_requests = paginator.page(paginator.num_pages)
    
    # Get distinct requesters for filter dropdown
    requesters = UsageLogs.objects.filter(approval_status="pending")\
        .values_list("requested_by__username", flat=True)\
        .distinct()
    
    context = {
        "pending_requests": pending_requests,
        "requesters": requesters,
        "search_query": search_query,
        "selected_requester": requester_filter,
    }
    return render(request, "approver/approver_dashboard.html", context)

@login_required
@role_required('approver')
def approval_history(request):
    """ Separate page for approval history """
    processed_requests = UsageLogs.objects.exclude(approval_status="pending")\
        .select_related('dsc_id', 'requested_by', 'approved_by_user')\
        .order_by('-date_of_usage')
    
    # Search and filters
    search_query = request.GET.get('search', '')
    requester_filter = request.GET.get('requester', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if search_query:
        processed_requests = processed_requests.filter(
            Q(dsc_id__dsc_number__icontains=search_query) |
            Q(dsc_id__full_name__icontains=search_query) |
            Q(form_name__icontains=search_query) |
            Q(platform_name__icontains=search_query)
        )    
    if requester_filter:
        processed_requests = processed_requests.filter(requested_by__username=requester_filter)
    
    if status_filter:
        processed_requests = processed_requests.filter(approval_status=status_filter)
    
    if date_from:
        processed_requests = processed_requests.filter(date_of_usage__gte=date_from)
    
    if date_to:
        processed_requests = processed_requests.filter(date_of_usage__lte=date_to)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(processed_requests, 15)
    
    try:
        processed_requests = paginator.page(page)
    except PageNotAnInteger:
        processed_requests = paginator.page(1)
    except EmptyPage:
        processed_requests = paginator.page(paginator.num_pages)
    
    # Get distinct requesters for filter dropdown
    requesters = UsageLogs.objects.exclude(approval_status="pending")\
        .values_list("requested_by__username", flat=True)\
        .distinct()
    
    context = {
        "processed_requests": processed_requests,
        "requesters": requesters,
        "search_query": search_query,
        "selected_requester": requester_filter,
        "selected_status": status_filter,
        "date_from": date_from,
        "date_to": date_to,
    }
    return render(request, "approver/approval_history.html", context)

# Keep your existing approve_usage_log and reject_usage_log views

@login_required
@role_required('approver')
def approve_usage_log(request, log_id):
    """ Approver approves a DSC usage log """
    usage_log = get_object_or_404(UsageLogs, log_id=log_id)
    
    if request.method == "POST":
        usage_log.approval_status = "approved"
        usage_log.approved_by_user = request.user
        usage_log.save()
        messages.success(request, f"✅ DSC request {log_id} approved!")
    
    return redirect("approver_dashboard")


@login_required
@role_required('approver')
def reject_usage_log(request, log_id):
    """ Approver rejects a DSC usage log """
    usage_log = get_object_or_404(UsageLogs, log_id=log_id)
    
    if request.method == "POST":
        usage_log.approval_status = "rejected"
        usage_log.approved_by_user = request.user
        usage_log.save()
        messages.warning(request, f"❌ DSC request {log_id} rejected.")
    
    return redirect("approver_dashboard")

##Backup Restore

@login_required
@role_required('admin')
def backup_db(request):
    db_path = settings.DATABASES['default']['NAME']
    if os.path.exists(db_path):
        temp = tempfile.NamedTemporaryFile(delete=False, suffix='.sqlite3')
        shutil.copy(db_path, temp.name)
        response = FileResponse(open(temp.name, 'rb'), as_attachment=True, filename='backup.sqlite3')
        return response
    else:
        messages.error(request, "❌ Database file not found.")
        return redirect('backup_restore/admin_dashboard')


@login_required
@role_required('admin')
def restore_db(request):
    if request.method == 'POST' and request.FILES.get('restore_file'):
        restore_file = request.FILES['restore_file']
        fs = FileSystemStorage(location=tempfile.gettempdir())
        filename = fs.save(restore_file.name, restore_file)
        uploaded_file_path = os.path.join(tempfile.gettempdir(), filename)

        db_path = settings.DATABASES['default']['NAME']
        shutil.copy(uploaded_file_path, db_path)

        messages.success(request, "✅ Database restored successfully.")
        return redirect('admin_dashboard')

    return render(request, 'backup_restore/admin_restore.html')

# @login_required
# @role_required('admin')
# def export_readable_backup(request):
#     """Export readable JSON version of the entire DB."""
#     buffer = io.StringIO()
#     call_command('dumpdata', indent=4, stdout=buffer)
#     buffer.seek(0)
    
#     response = HttpResponse(buffer, content_type='application/json')
#     timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
#     response['Content-Disposition'] = f'attachment; filename=readable_backup_{timestamp}.json'
#     return response

# @login_required
# @role_required('admin')
# def export_full_excel_backup(request):
#     wb = Workbook()

#     # 1. DSCMaster Sheet
#     ws1 = wb.active
#     ws1.title = "DSCs"
#     ws1.append(['DSC Number', 'Full Name', 'Issued Date', 'Expiry Date', 'Status', 'Email', 'Phone'])

#     for d in DSCMaster.objects.all():
#         ws1.append([
#             d.dsc_number, d.full_name,
#             d.issued_date.strftime("%Y-%m-%d") if d.issued_date else '',
#             d.expiry_date.strftime("%Y-%m-%d") if d.expiry_date else '',
#             d.status, d.email_id, d.phone_no
#         ])

#     # 2. Usage Logs Sheet
#     ws2 = wb.create_sheet("Usage Logs")
#     ws2.append(['DSC Number', 'Form', 'Platform', 'Date', 'Status'])

#     for u in UsageLogs.objects.all():
#         ws2.append([
#             u.dsc_id.dsc_number if u.dsc_id else '',
#             u.form_name,
#             u.platform_name,
#             u.date_of_usage.strftime('%Y-%m-%d'),
#             u.approval_status
#         ])

#     # 3. Entities
#     ws3 = wb.create_sheet("Entities")
#     ws3.append(['Entity ID', 'Entity Name'])
#     for e in Entity.objects.all():
#         ws3.append([e.entity_id, e.entity_name])

#     # 4. Platforms
#     ws4 = wb.create_sheet("Platforms")
#     ws4.append(['Platform ID', 'Platform Name'])
#     for p in Platform.objects.all():
#         ws4.append([p.platform_id, p.platform_name])

#     # 5. Forms
#     ws5 = wb.create_sheet("Forms")
#     ws5.append(['Form ID', 'Form Name'])
#     for f in Forms.objects.all():
#         ws5.append([f.form_id, f.form_name])

#     # 6. DSCRenewals
#     # ws6 = wb.create_sheet("DSC Renewals")
#     # ws6.append(['DSC Number', 'Remarks', 'Timestamp'])
#     # for r in DSCRenewal.objects.all():
#     #     ws6.append([r.dsc_number.dsc_number, r.remarks, r.timestamp.strftime("%Y-%m-%d %H:%M:%S")])

#     # Response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     filename = f"full_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
#     response['Content-Disposition'] = f'attachment; filename={filename}'
#     wb.save(response)
#     return response


@login_required
@role_required('admin')
def export_all_models_excel(request):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    model_names = [
        'CustomUser', 'DSC_class', 'Platform', 'IssuingAuth', 'Entity',
        'LicensePeriod', 'Forms', 'Type', 'Shelf',
        'DSCMaster', 'DSCEntity', 'Docs', 'InOut', 'DSCRenewal', 'UsageLogs',
        'DSCExtraField', 'DSCExtraData', 'EmailTemplate'
    ]

    app_label = 'home'  # 🔁 Replace this with your actual app name

    for model_name in model_names:
        model = apps.get_model(app_label, model_name)
        objects = model.objects.all()

        # Special handling for DSCMaster foreign keys
        if model_name == 'DSCMaster':
            data = []
            for obj in objects:
                data.append({
                    'DSC_number': obj.dsc_number,
                    'Full Name': obj.full_name,
                    'Type': obj.dsc_type.type_name if obj.dsc_type else None,
                    'Class': obj.dsc_class.class_name if obj.dsc_class else None,
                    'Issued Date': obj.issued_date,
                    'Expiry Date': obj.expiry_date,
                    'License Period': f"{obj.license_period.no_of_years} years" if obj.license_period else None,
                    'Issuing Authority': obj.issuing_auth.auth_name if obj.issuing_auth else None,
                    'Email': obj.email_id,
                    'Phone': obj.phone_no,
                    'Reference Name': obj.ref_name,
                    'Reference Contact': obj.ref_contact,
                    'Organization/Individual': obj.type,
                    'Status': obj.status,
                    'User': str(obj.user_id) if obj.user_id else None,
                    'Remarks': obj.remarks,
                    'Extra fields': obj.extra_fields if obj.extra_fields else None,
                })
            df = pd.DataFrame(data)

        elif model_name == 'DSCEntity':
            data = []
            for obj in objects:
                data.append({
                    'DSC Number': obj.dsc_number,
                    'Entity':obj.entity,
                })
            df = pd.DataFrame(data)

        elif model_name == 'Docs':
            data = []
            for obj in objects:
                data.append({
                    'Docs id': obj.docs_id,
                    'DSC number': obj.dsc_number,
                    'Aadhar path': obj.aadhar_path if obj.aadhar_path else None,
                    'Pan path': obj.pan_path if obj.pan_path else None,
                    'PP path': obj.pp_path if obj.pp_path else None,
                })
            df = pd.DataFrame(data)
            
        elif model_name == 'InOut':
            data = []
            for obj in objects:
                data.append({
                    'INOUT id': obj.in_out_id,
                    'DSC_number': obj.dsc_number,
                    'Direction': obj.direction if obj.direction else None,
                    'Initiated By': obj.initiated_by if obj.initiated_by else None,
                    'Initiation Date': obj.initiation_date.strftime("%Y-%m-%d") if obj.initiation_date else None,
                    'Requester Name': obj.requester_name if obj.requester_name else None,
                    'Initiation Remarks': obj.initiation_remarks if obj.initiation_remarks else '',
                    'Initiates': obj.initiated,
                    'Action Date': obj.action_date.strftime("%Y-%m-%d") if obj.action_date else '',
                    'Action Time': obj.action_time if obj.action_time else None,
                    'Agent Name': obj.agent_name if obj.agent_name else None,
                    'Agent Mob': obj.agent_mob if obj.agent_mob else None,
                    'Action Remarks': obj.action_remarks if obj.action_remarks else None,
                    'Action Completed': obj.action_completed

                })
            df = pd.DataFrame(data)

        elif model_name == 'DSCRenewal':
            data = []
            for obj in objects:
                data.append({
                    'Id': obj.additional_id,
                    'DSC number': obj.dsc_number,
                    'GST reg date': obj.gst_reg_date,
                    'IT reg date': obj.it_reg_date,
                    'MCA reg date': obj.mca_reg_date,
                    'Remarks': obj.remarks,
                })
            df = pd.DataFrame(data)      

        elif model_name == 'UsageLogs':
            data = []
            for obj in objects:
                data.append({
                    'Log Id': obj.log_id,
                    'DSC number': obj.dsc_id,
                    'Form Id': obj.form_id,
                    'Form Other': obj.form_other,
                    'Platform Id': obj.platform_id,
                    'Platform Other': obj.platform_other,
                    'Date Of Usage': obj.date_of_usage,
                    'Requested By': obj.requested_by,
                    'Approved By User': obj.approved_by_user,
                    'Approval Status': obj.approval_status,
                })
            df = pd.DataFrame(data)   

        elif model_name == 'DSCExtraData':
            data = []
            for obj in objects:
                data.append({
                    'DSC Number': obj.dsc_number,
                    'Field Name':obj.field_name,
                    'Field Value':obj.value,
                })
            df = pd.DataFrame(data)                    

        else:
            df = pd.DataFrame(list(objects.values()))
            # Convert timezone-aware datetimes to naive
            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col].apply(lambda x: x.replace(tzinfo=None) if pd.notnull(x) and hasattr(x, 'tzinfo') else x)

        # Write each model data to a separate sheet
        df.to_excel(writer, sheet_name=model_name[:31], index=False)

    writer.close()
    output.seek(0)

    filename = f"DSC_Backup_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


##modification

from django.db.models import Max

@login_required
@role_required('admin')
def shelf_history(request):
    # Get all assignment/unassignment records
    history = ShelfAssignment.objects.select_related(
        'dsc_number', 
        'shelf_no',
        'action_by'
    ).order_by('-action_date')

    # Initialize filter variables
    dsc_number = request.GET.get('dsc_number', '')
    shelf_no = request.GET.get('shelf_no', '')
    performed_by = request.GET.get('performed_by', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    action_type = request.GET.get('action_type', '')

    # Apply filters
    if dsc_number:
        history = history.filter(dsc_number__dsc_number__icontains=dsc_number)
    if shelf_no:
        history = history.filter(shelf_no__shelf_no__icontains=shelf_no)
    if performed_by:
        history = history.filter(action_by__username__icontains=performed_by)
    if action_type:
        history = history.filter(action=action_type)
    if start_date:
        history = history.filter(action_date__gte=start_date)
    if end_date:
        history = history.filter(action_date__lte=end_date)

    # Pagination
    paginator = Paginator(history, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_records': history.count(),
        'filter_values': {
            'dsc_number': dsc_number,
            'shelf_no': shelf_no,
            'performed_by': performed_by,
            'start_date': start_date,
            'end_date': end_date,
            'action_type': action_type
        }
    }
    return render(request, 'shelf_history.html', context)

@login_required
@role_required('admin')
def assign_shelf(request):
    # Get all DSCs that are 'in' and not currently assigned
    available_dscs = []
    all_dscs = DSCMaster.objects.select_related('dsc_type').prefetch_related('inout_set').all()
    
    # Get most recent action for each DSC
    latest_actions = ShelfAssignment.objects.values('dsc_number').annotate(
        latest_date=Max('action_date')
    ).order_by()
    
    # Create dictionary of latest action types
    latest_action_types = {}
    for item in latest_actions:
        action = ShelfAssignment.objects.filter(
            dsc_number=item['dsc_number'],
            action_date=item['latest_date']
        ).first()
        latest_action_types[action.dsc_number_id] = action.action
    
    for dsc in all_dscs:
        # Check direction
        is_in = dsc.current_direction == 'in'
        
        # Check if DSC has any assignment history
        has_history = dsc.dsc_id in latest_action_types
        
        # DSC is available if:
        # 1. It's 'in' AND
        # 2. Either has no history OR last action was 'unassign'
        if is_in and (not has_history or latest_action_types[dsc.dsc_id] == 'unassign'):
            available_dscs.append(dsc)

    if request.method == 'POST':
        form = AssignShelfForm(request.POST)
        if form.is_valid():
            dsc_number = request.POST.get('dsc_number')
            try:
                dsc = DSCMaster.objects.get(dsc_number=dsc_number)
                
                # Re-validate using model property
                if dsc.current_direction != 'in':
                    messages.error(request, f"Cannot assign shelf - DSC {dsc.dsc_number} is currently 'out'")
                    return redirect('assign_shelf')
                
                # Check if already assigned (last action was 'assign')
                last_action = ShelfAssignment.objects.filter(
                    dsc_number=dsc
                ).order_by('-action_date').first()
                
                if last_action and last_action.action == 'assign':
                    messages.error(request, f"DSC {dsc.dsc_number} already has a shelf assigned")
                    return redirect('assign_shelf')

                assignment = form.save(commit=False)
                assignment.dsc_number = dsc
                assignment.action = 'assign'
                assignment.action_by = request.user
                assignment.save()
                
                messages.success(request, f"Shelf {assignment.shelf_no} assigned to DSC {dsc.dsc_number}")
                return redirect('assign_shelf')
            except DSCMaster.DoesNotExist:
                messages.error(request, "Selected DSC does not exist")
                return redirect('assign_shelf')
    else:
        form = AssignShelfForm()

    context = {
        'available_dscs': available_dscs,
        'form': form,
        'shelves': Shelf.objects.all(),
        'dsc_list': available_dscs
    }
    return render(request, 'assign_shelf.html', context)

@login_required
@role_required('admin')
def unassign_shelf(request):
    # Get all DSCs where most recent action is 'assign'
    assigned_dscs = []
    
    # Get most recent action for each DSC
    latest_actions = ShelfAssignment.objects.values('dsc_number').annotate(
        latest_date=Max('action_date')
    ).order_by()
    
    # Filter to only include DSCs where latest action is 'assign'
    active_assignments = ShelfAssignment.objects.filter(
        action='assign',
        action_date__in=[item['latest_date'] for item in latest_actions]
    ).select_related('dsc_number', 'shelf_no', 'action_by')
    
    for assignment in active_assignments:
        assigned_dscs.append({
            'dsc': assignment.dsc_number,
            'shelf': assignment.shelf_no,
            'assignment_date': assignment.action_date,
            'assigned_by': assignment.action_by,
        })

    if request.method == 'POST':
        form = UnassignShelfForm(request.POST)
        if form.is_valid():
            dsc_number = request.POST.get('dsc_number')
            try:
                dsc = DSCMaster.objects.get(dsc_number=dsc_number.split(' - ')[0])
                
                # Get the current active assignment (most recent 'assign' action)
                current_assignment = ShelfAssignment.objects.filter(
                    dsc_number=dsc,
                    action='assign'
                ).order_by('-action_date').first()

                if not current_assignment:
                    messages.error(request, f"No active shelf assignment found for DSC {dsc.dsc_number}")
                    return redirect('unassign_shelf')

                # Create unassignment record (DO NOT DELETE THE ASSIGNMENT RECORD)
                ShelfAssignment.objects.create(
                    dsc_number=dsc,
                    shelf_no=current_assignment.shelf_no,
                    action='unassign',
                    action_by=request.user,
                    remarks=form.cleaned_data['remarks']
                )
                
                messages.success(request, 
                    f"Shelf {current_assignment.shelf_no} unassigned from DSC {dsc.dsc_number}")
                return redirect('unassign_shelf')
            except DSCMaster.DoesNotExist:
                messages.error(request, "Selected DSC does not exist")
                return redirect('unassign_shelf')
    else:
        form = UnassignShelfForm()

    context = {
        'assigned_dscs': assigned_dscs,
        'form': form,
        'dsc_list': [item['dsc'] for item in assigned_dscs]
    }
    return render(request, 'unassign_shelf.html', context)



@login_required
@role_required('general-user')
def general_user_shelf_history(request):
    # Get all assignment/unassignment records
    history = ShelfAssignment.objects.select_related(
        'dsc_number', 
        'shelf_no',
        'action_by'
    ).order_by('-action_date')

    # Initialize filter variables
    dsc_number = request.GET.get('dsc_number', '')
    shelf_no = request.GET.get('shelf_no', '')
    performed_by = request.GET.get('performed_by', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    action_type = request.GET.get('action_type', '')

    # Apply filters
    if dsc_number:
        history = history.filter(dsc_number__dsc_number__icontains=dsc_number)
    if shelf_no:
        history = history.filter(shelf_no__shelf_no__icontains=shelf_no)
    if performed_by:
        history = history.filter(action_by__username__icontains=performed_by)
    if action_type:
        history = history.filter(action=action_type)
    if start_date:
        history = history.filter(action_date__gte=start_date)
    if end_date:
        history = history.filter(action_date__lte=end_date)

    # Pagination
    paginator = Paginator(history, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_records': history.count(),
        'filter_values': {
            'dsc_number': dsc_number,
            'shelf_no': shelf_no,
            'performed_by': performed_by,
            'start_date': start_date,
            'end_date': end_date,
            'action_type': action_type
        }
    }
    return render(request, 'general_user/shelf_history.html', context)

@login_required
@role_required('general-user')
def general_user_assign_shelf(request):
    # Get all DSCs that are 'in' and not currently assigned
    available_dscs = []
    all_dscs = DSCMaster.objects.select_related('dsc_type').prefetch_related('inout_set').all()
    
    # Get most recent action for each DSC
    latest_actions = ShelfAssignment.objects.values('dsc_number').annotate(
        latest_date=Max('action_date')
    ).order_by()
    
    # Create dictionary of latest action types
    latest_action_types = {}
    for item in latest_actions:
        action = ShelfAssignment.objects.filter(
            dsc_number=item['dsc_number'],
            action_date=item['latest_date']
        ).first()
        latest_action_types[action.dsc_number_id] = action.action
    
    for dsc in all_dscs:
        # Check direction
        is_in = dsc.current_direction == 'in'
        
        # Check if DSC has any assignment history
        has_history = dsc.dsc_id in latest_action_types
        
        # DSC is available if:
        # 1. It's 'in' AND
        # 2. Either has no history OR last action was 'unassign'
        if is_in and (not has_history or latest_action_types[dsc.dsc_id] == 'unassign'):
            available_dscs.append(dsc)

    if request.method == 'POST':
        form = AssignShelfForm(request.POST)
        if form.is_valid():
            dsc_number = request.POST.get('dsc_number')
            try:
                dsc = DSCMaster.objects.get(dsc_number=dsc_number)
                
                # Re-validate using model property
                if dsc.current_direction != 'in':
                    messages.error(request, f"Cannot assign shelf - DSC {dsc.dsc_number} is currently 'out'")
                    return redirect('general_user_assign_shelf')
                
                # Check if already assigned (last action was 'assign')
                last_action = ShelfAssignment.objects.filter(
                    dsc_number=dsc
                ).order_by('-action_date').first()
                
                if last_action and last_action.action == 'assign':
                    messages.error(request, f"DSC {dsc.dsc_number} already has a shelf assigned")
                    return redirect('general_user_assign_shelf')

                assignment = form.save(commit=False)
                assignment.dsc_number = dsc
                assignment.action = 'assign'
                assignment.action_by = request.user
                assignment.save()
                
                messages.success(request, f"Shelf {assignment.shelf_no} assigned to DSC {dsc.dsc_number}")
                return redirect('general_user_assign_shelf')
            except DSCMaster.DoesNotExist:
                messages.error(request, "Selected DSC does not exist")
                return redirect('general_user_assign_shelf')
    else:
        form = AssignShelfForm()

    context = {
        'available_dscs': available_dscs,
        'form': form,
        'shelves': Shelf.objects.all(),
        'dsc_list': available_dscs
    }
    return render(request, 'general_user/assign_shelf.html', context)

@login_required
@role_required('general-user')
def general_user_unassign_shelf(request):
    # Get all DSCs where most recent action is 'assign'
    assigned_dscs = []
    
    # Get most recent action for each DSC
    latest_actions = ShelfAssignment.objects.values('dsc_number').annotate(
        latest_date=Max('action_date')
    ).order_by()
    
    # Filter to only include DSCs where latest action is 'assign'
    active_assignments = ShelfAssignment.objects.filter(
        action='assign',
        action_date__in=[item['latest_date'] for item in latest_actions]
    ).select_related('dsc_number', 'shelf_no', 'action_by')
    
    for assignment in active_assignments:
        assigned_dscs.append({
            'dsc': assignment.dsc_number,
            'shelf': assignment.shelf_no,
            'assignment_date': assignment.action_date,
            'assigned_by': assignment.action_by,
        })

    if request.method == 'POST':
        form = UnassignShelfForm(request.POST)
        if form.is_valid():
            dsc_number = request.POST.get('dsc_number')
            try:
                dsc = DSCMaster.objects.get(dsc_number=dsc_number.split(' - ')[0])
                
                # Get the current active assignment (most recent 'assign' action)
                current_assignment = ShelfAssignment.objects.filter(
                    dsc_number=dsc,
                    action='assign'
                ).order_by('-action_date').first()

                if not current_assignment:
                    messages.error(request, f"No active shelf assignment found for DSC {dsc.dsc_number}")
                    return redirect('general_user_unassign_shelf')

                # Create unassignment record (DO NOT DELETE THE ASSIGNMENT RECORD)
                ShelfAssignment.objects.create(
                    dsc_number=dsc,
                    shelf_no=current_assignment.shelf_no,
                    action='unassign',
                    action_by=request.user,
                    remarks=form.cleaned_data['remarks']
                )
                
                messages.success(request, 
                    f"Shelf {current_assignment.shelf_no} unassigned from DSC {dsc.dsc_number}")
                return redirect('general_user_unassign_shelf')
            except DSCMaster.DoesNotExist:
                messages.error(request, "Selected DSC does not exist")
                return redirect('general_user_unassign_shelf')
    else:
        form = UnassignShelfForm()

    context = {
        'assigned_dscs': assigned_dscs,
        'form': form,
        'dsc_list': [item['dsc'] for item in assigned_dscs]
    }
    return render(request, 'general_user/unassign_shelf.html', context)


from django.db import transaction
from io import StringIO
import pandas as pd
from datetime import datetime

@login_required
@role_required('admin')
def bulk_upload_dsc(request):
    if request.method == 'POST':
        form = BulkDSCUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                dsc_type = form.cleaned_data['dsc_type']
                upload_file = request.FILES['upload_file']
                
                # Read the uploaded file
                if upload_file.name.endswith('.csv'):
                    df = pd.read_csv(upload_file)
                else:  # Excel file
                    df = pd.read_excel(upload_file)
                
                # Convert to list of dictionaries
                records = df.replace({pd.NA: None}).to_dict('records')
                
                created_count = 0
                errors = []
                
                with transaction.atomic():
                    for idx, record in enumerate(records, start=2):  # Start at 2 for Excel row numbers
                        try:
                            if dsc_type == 'internal':
                                dsc = process_internal_dsc_record(record, request.user)
                            else:
                                dsc = process_external_dsc_record(record, request.user)
                            
                            created_count += 1
                        except Exception as e:
                            errors.append(f"Row {idx}: {str(e)}")
                
                if errors:
                    messages.warning(
                        request, 
                        f"Successfully created {created_count} DSCs with {len(errors)} errors"
                    )
                    return render(request, 'bulk_upload_errors.html', {'errors': errors})
                else:
                    messages.success(request, f"Successfully created {created_count} DSCs")
                    return redirect('map_dsc_entity')
                    
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = BulkDSCUploadForm()
    
    return render(request, 'bulk_upload_dsc.html', {'form': form})

def process_internal_dsc_record(record, user):
    """Process a single internal DSC record from bulk upload with all required fields"""
    # All fields that are required in InternalDSCForm
    required_fields = [
        'dsc_number', 'full_name', 'issued_date', 'license_period', 'password',
        'pan_no', 'dsc_class', 'email_id', 'phone_no', 'issuing_auth',
        'ref_name', 'ref_contact', 'type', 'remarks'
    ]
    
    # Validate required fields
    for field in required_fields:
        if field not in record or not record[field]:
            raise ValidationError(f"Missing required field for internal DSC: {field}")
    
    # Auto-generate DSC number if not provided (though InternalDSCForm makes it required)
    if not record.get('dsc_number'):
        internal_type = Type.objects.get(type_name='Internal')
        last_internal = DSCMaster.objects.filter(dsc_type=internal_type).order_by('-dsc_id').first()
        dsc_number = str(int(last_internal.dsc_number) + 1) if last_internal else "1500"
    else:
        dsc_number = record['dsc_number']
    
    # Convert date strings to date objects
    if isinstance(record['issued_date'], str):
        try:
            record['issued_date'] = datetime.strptime(record['issued_date'], '%Y-%m-%d').date()
        except ValueError:
            raise ValidationError("Invalid date format. Use YYYY-MM-DD")
    
    # Get license period object
    try:
        if isinstance(record['license_period'], str):
            license_period = LicensePeriod.objects.get(no_of_years=int(record['license_period']))
        else:
            license_period = LicensePeriod.objects.get(pk=record['license_period'])
    except (LicensePeriod.DoesNotExist, ValueError):
        raise ValidationError("Invalid license period")
    
    # Create DSC with all required fields
    dsc = DSCMaster(
        dsc_number=dsc_number,
        full_name=record['full_name'],
        issued_date=record['issued_date'],
        license_period=license_period,
        dsc_type=Type.objects.get(type_name='Internal'),
        user_id=user,
        pan_no=record['pan_no'],
        email_id=record['email_id'],
        phone_no=record['phone_no'],
        dsc_class_id=record['dsc_class'],
        issuing_auth_id=record['issuing_auth'],
        ref_name=record['ref_name'],
        ref_contact=record['ref_contact'],
        type=record['type'],
        remarks=record['remarks'],
        password=record['password']
    )
    
    dsc.full_clean()
    dsc.save()
    
    # Create documents record (required for internal)
    Docs.objects.create(dsc_number=dsc)
    
    # Create InOut record
    InOut.objects.create(
        dsc_number=dsc,
        direction="in",
        initiated=False,
        action_completed=False
    )
    
    return dsc

def process_external_dsc_record(record, user):
    """Process a single external DSC record from bulk upload with proper null handling"""
    # Required fields for external DSC
    required_fields = ['dsc_number', 'full_name', 'issued_date', 'license_period', 'password', 'remarks']
    
    # Validate required fields
    for field in required_fields:
        if field not in record or pd.isna(record[field]) or str(record[field]).strip() == '':
            raise ValidationError(f"Missing required field for external DSC: {field}")
    
    # Convert date strings to date objects
    if isinstance(record['issued_date'], str):
        try:
            record['issued_date'] = datetime.strptime(record['issued_date'], '%Y-%m-%d').date()
        except ValueError:
            raise ValidationError("Invalid date format. Use YYYY-MM-DD")
    
    # Get license period object
    try:
        if isinstance(record['license_period'], str):
            license_period = LicensePeriod.objects.get(no_of_years=int(float(record['license_period'])))
        else:
            license_period = LicensePeriod.objects.get(pk=int(float(record['license_period'])))
    except (LicensePeriod.DoesNotExist, ValueError, TypeError):
        raise ValidationError("Invalid license period")
    
    # Helper function to clean optional foreign key fields
    def clean_foreign_key(value):
        if pd.isna(value) or str(value).strip() == '':
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None
    
    # Helper function to clean optional choice fields
    def clean_choice_field(value, choices):
        if pd.isna(value) or str(value).strip() == '':
            return None
        value = str(value).strip()
        if value in [choice[0] for choice in choices]:
            return value
        return None
    
    # Create DSC instance
    dsc = DSCMaster(
        dsc_number=str(record['dsc_number']).strip(),
        full_name=str(record['full_name']).strip(),
        issued_date=record['issued_date'],
        license_period=license_period,
        dsc_type=Type.objects.get(type_name='External'),
        user_id=user,
        password=str(record['password']).strip(),
        remarks=str(record['remarks']).strip(),
        # Optional fields
        pan_no=clean_optional_field(record.get('pan_no')),
        email_id=clean_optional_field(record.get('email_id')),
        phone_no=clean_optional_field(record.get('phone_no')),
        dsc_class_id=clean_foreign_key(record.get('dsc_class')),
        issuing_auth_id=clean_foreign_key(record.get('issuing_auth')),
        ref_name=clean_optional_field(record.get('ref_name')),
        ref_contact=clean_optional_field(record.get('ref_contact')),
        type=clean_choice_field(record.get('type'), 
                              [('Organization', 'Organization'), ('Individual', 'Individual')])
    )
    
    # Additional validation for cleaned data
    if dsc.email_id and not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', dsc.email_id):
        dsc.email_id = None
    
    if dsc.phone_no and not re.match(r'^\d{10}$', dsc.phone_no):
        dsc.phone_no = None
    
    if dsc.ref_contact and not re.match(r'^\d{10}$', dsc.ref_contact):
        dsc.ref_contact = None
    
    dsc.full_clean()
    dsc.save()
    
    # Create documents record
    Docs.objects.create(dsc_number=dsc)
    
    return dsc

def clean_optional_field(value):
    """Clean general optional fields"""
    if pd.isna(value) or str(value).strip() == '':
        return None
    return str(value).strip()

@login_required
@role_required('admin')
def download_bulk_template(request):
    dsc_type = request.GET.get('type', 'internal')
    
    if dsc_type == 'internal':
        # Internal template with all required fields
        data = {
            'dsc_number': ['INT-001'],  # Will be auto-generated if empty
            'full_name': ['John Doe'],
            'issued_date': ['2023-01-01'],
            'license_period': ['1'],  # Years or ID
            'password': ['SecurePass123!'],
            'pan_no': ['ABCDE1234F'],
            'email_id': ['john@example.com'],
            'phone_no': ['9876543210'],
            'dsc_class': ['1'],  # IDs from DSC_class
            'issuing_auth': ['1'],  # IDs from IssuingAuth
            'ref_name': ['Ref Person'],
            'ref_contact': ['9876543210'],
            'type': ['Individual'],
            'remarks': ['Sample remark']
        }
    else:
        # External template with only required fields
        data = {
            'dsc_number': ['EXT-001'],
            'full_name': ['External User'],
            'issued_date': ['2023-01-01'],
            'license_period': ['1'],  # Years or ID
            'password': ['SecurePass123!'],
            'remarks': ['Sample remark'],
            # Optional fields with empty values
            'pan_no': [''],
            'email_id': [''],
            'phone_no': [''],
            'dsc_class': [''],
            'issuing_auth': [''],
            'ref_name': [''],
            'ref_contact': [''],
            'type': ['']
        }
    
    df = pd.DataFrame(data)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="dsc_bulk_template_{dsc_type}.xlsx"'
    
    df.to_excel(response, index=False)
    return response


@login_required
@role_required('admin')
def bulk_upload_entity(request):
    if request.method == 'POST':
        form = BulkEntityUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                upload_file = request.FILES['upload_file']
                
                # Read the uploaded file
                if upload_file.name.endswith('.csv'):
                    df = pd.read_csv(upload_file)
                else:  # Excel file
                    df = pd.read_excel(upload_file)
                
                # Convert to list of dictionaries
                records = df.replace({pd.NA: None}).to_dict('records')
                
                created_count = 0
                errors = []
                
                with transaction.atomic():
                    for idx, record in enumerate(records, start=2):  # Start at 2 for Excel row numbers
                        try:
                            entity = process_entity_record(record)
                            created_count += 1
                        except Exception as e:
                            errors.append(f"Row {idx}: {str(e)}")
                
                if errors:
                    messages.warning(
                        request, 
                        f"Successfully created {created_count} entities with {len(errors)} errors"
                    )
                    return render(request, 'bulk_upload_errors.html', {'errors': errors})
                else:
                    messages.success(request, f"Successfully created {created_count} entities")
                    return redirect('entity_list')  # Replace with your entity list view
                    
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = BulkEntityUploadForm()
    
    return render(request, 'bulk_upload_entity.html', {'form': form})

def process_entity_record(record):
    """Process a single entity record from bulk upload"""
    if 'entity_name' not in record or not record['entity_name']:
        raise ValidationError("Missing required field: entity_name")
    
    entity_name = str(record['entity_name']).strip()
    
    # Check for duplicates
    if Entity.objects.filter(entity_name__iexact=entity_name).exists():
        raise ValidationError(f"Entity with name '{entity_name}' already exists")
    
    entity = Entity(entity_name=entity_name)
    entity.full_clean()
    entity.save()
    
    return entity

@login_required
@role_required('admin')
def download_entity_template(request):
    data = {
        'entity_name': ['Entity 1', 'Entity 2', 'Entity 3']
    }
    
    df = pd.DataFrame(data)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="entity_bulk_template.xlsx"'
    
    df.to_excel(response, index=False)
    return response


# views.py

@login_required
@role_required('admin')
def renewal_history(request, dsc_id):
    dsc = get_object_or_404(DSCMaster, pk=dsc_id)
    history_entries = dsc.renewal_history.all().order_by('-renewal_date')
    
    # Prepare data for display
    history_data = []
    for entry in history_entries:
        history_data.append({
            'renewal_date': entry.renewal_date,
            'renewed_by': entry.renewed_by,
            'changes': get_field_changes(entry.previous_data, entry.new_data),
            'previous_documents': entry.previous_documents,  # Changed from documents_snapshot
            'new_documents': entry.new_documents,            # Changed from documents_snapshot
            'document_changes': entry.document_changes       # Add this line
        })
    
    return render(request, 'renewal_history.html', {
        'dsc': dsc,
        'history': history_data,
        'current_dsc_data': {
            'issued_date': dsc.issued_date,
            'expiry_date': dsc.expiry_date,
            'status': dsc.status,
            'dsc_number': dsc.dsc_number,
            'full_name': dsc.full_name,
        }
    })

def get_field_changes(old_data, new_data):
    changes = []
    all_fields = set(old_data.keys()).union(set(new_data.keys()))
    
    for field in all_fields:
        old_value = old_data.get(field)
        new_value = new_data.get(field)
        
        if old_value != new_value:
            # Format dates properly
            if field.endswith('_date') and old_value:
                try:
                    old_value = datetime.strptime(old_value, '%Y-%m-%d').strftime('%d %b %Y')
                except:
                    pass
            if field.endswith('_date') and new_value:
                try:
                    new_value = datetime.strptime(new_value, '%Y-%m-%d').strftime('%d %b %Y')
                except:
                    pass
            
            # Handle special cases
            if field == 'license_period':
                try:
                    old_value = LicensePeriod.objects.get(license_id=old_value).no_of_years if old_value else None
                    new_value = LicensePeriod.objects.get(license_id=new_value).no_of_years if new_value else None
                except LicensePeriod.DoesNotExist:
                    pass
            
            changes.append({
                'field': field.replace('_', ' ').title(),
                'old': old_value,
                'new': new_value
            })
    
    return changes

@login_required
@role_required('general-user')
def general_user_renewal_history(request, dsc_id):
    dsc = get_object_or_404(DSCMaster, pk=dsc_id)
    history_entries = dsc.renewal_history.all().order_by('-renewal_date')
    
    # Prepare data for display
    history_data = []
    for entry in history_entries:
        history_data.append({
            'renewal_date': entry.renewal_date,
            'renewed_by': entry.renewed_by,
            'changes': get_field_changes(entry.previous_data, entry.new_data),
            'previous_documents': entry.previous_documents,  # Changed from documents_snapshot
            'new_documents': entry.new_documents,            # Changed from documents_snapshot
            'document_changes': entry.document_changes       # Add this line
        })
    
    return render(request, 'general_user/renewal_history.html', {
        'dsc': dsc,
        'history': history_data,
        'current_dsc_data': {
            'issued_date': dsc.issued_date,
            'expiry_date': dsc.expiry_date,
            'status': dsc.status,
            'dsc_number': dsc.dsc_number,
            'full_name': dsc.full_name,
        }
    })

from django.db.models import Q
from datetime import datetime

@login_required
@role_required('admin')
def all_renewals(request):
    # Get all renewal history entries
    renewals = DSCRenewalHistory.objects.all().order_by('-renewal_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        renewals = renewals.filter(
            Q(dsc__dsc_number__icontains=search_query) |
            Q(dsc__full_name__icontains=search_query) |
            Q(renewed_by__username__icontains=search_query)
        )
    
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            renewals = renewals.filter(renewal_date__gte=start_date)
        except ValueError:
            messages.error(request, "Invalid start date format. Use YYYY-MM-DD.")
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            renewals = renewals.filter(renewal_date__lte=end_date)
        except ValueError:
            messages.error(request, "Invalid end date format. Use YYYY-MM-DD.")
    
    # Pagination
    paginator = Paginator(renewals, 25)  # Show 25 renewals per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get current filter values for the template
    current_filters = {
        'search': search_query,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
    }
    
    return render(request, 'all_renewals.html', {
        'page_obj': page_obj,
        'total_renewals': renewals.count(),
        'current_filters': current_filters
    })

@login_required
@role_required('general-user')
def general_user_all_renewals(request):
    # Get all renewal history entries
    renewals = DSCRenewalHistory.objects.all().order_by('-renewal_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        renewals = renewals.filter(
            Q(dsc__dsc_number__icontains=search_query) |
            Q(dsc__full_name__icontains=search_query) |
            Q(renewed_by__username__icontains=search_query)
        )
    
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            renewals = renewals.filter(renewal_date__gte=start_date)
        except ValueError:
            messages.error(request, "Invalid start date format. Use YYYY-MM-DD.")
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            renewals = renewals.filter(renewal_date__lte=end_date)
        except ValueError:
            messages.error(request, "Invalid end date format. Use YYYY-MM-DD.")
    
    # Pagination
    paginator = Paginator(renewals, 25)  # Show 25 renewals per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get current filter values for the template
    current_filters = {
        'search': search_query,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
    }
    
    return render(request, 'general_user/all_renewals.html', {
        'page_obj': page_obj,
        'total_renewals': renewals.count(),
        'current_filters': current_filters
    })